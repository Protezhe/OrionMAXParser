import argparse
import datetime
import json
import io
import sys
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Color, Fill, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

# Импортируем matplotlib для построения крутой графики
import matplotlib
matplotlib.use('Agg')  # Фоновый режим без открытия окон
import matplotlib.pyplot as plt
from matplotlib.colors import to_rgb
from matplotlib.patches import Polygon, Rectangle
from matplotlib.backends.backend_pdf import PdfPages

# Константы цветов для таблиц (VBA репликация)
RGB_ORANGE = "FFC000"
RGB_YELLOW = "FFFF00"
RGB_GREEN = "70AD47"
RGB_BLUE = "00B0F0"
RGB_HEADER_BLUE = "D9EAF7"

# Цвета для графиков как в ведомости Excel
CHART_COLOR_DOWNTIME = "#ED7D31"  # Оранжевый для времени остановок
CHART_COLOR_COUNT = "#5B9BD5"     # Синий для количества остановок

def shade_color(color: str, factor: float) -> Tuple[float, float, float]:
    r, g, b = to_rgb(color)
    return (
        max(0, min(1, r * factor)),
        max(0, min(1, g * factor)),
        max(0, min(1, b * factor)),
    )

def load_config(config_path: Path) -> Dict[str, Any]:
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)

def get_month_name_ru(month: int) -> str:
    months = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    ]
    return months[month - 1]

def parse_time(value: Any) -> datetime.time | None:
    if isinstance(value, datetime.time):
        return value
    if isinstance(value, datetime.datetime):
        return value.time()
    if isinstance(value, str):
        try:
            return datetime.datetime.strptime(value, "%H:%M").time()
        except ValueError:
            return None
    return None

def time_to_seconds(t: datetime.time) -> int:
    return t.hour * 3600 + t.minute * 60 + t.second

def seconds_to_excel_time(seconds: float) -> float:
    return seconds / (24 * 3600)

def calculate_downtime(start: datetime.time, end: datetime.time) -> float:
    s = time_to_seconds(s) if isinstance(s := start, datetime.time) else time_to_seconds(parse_time(start))
    e = time_to_seconds(e) if isinstance(e := end, datetime.time) else time_to_seconds(parse_time(end))
    if e < s:
        diff = (24 * 3600 - s) + e
    else:
        diff = e - s
    return seconds_to_excel_time(diff)

def apply_border(cell):
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def draw_modern_chart(ax, title: str, categories: List[str], values: List[float], color: str, is_time: bool = False):
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['font.sans-serif'] = ['Arial', 'Calibri', 'DejaVu Sans']
    
    if is_time:
        plot_values = [v * 24 for v in values]
        ylabel = "Часы простоя"
    else:
        plot_values = values
        ylabel = "Количество остановок"
        
    bar_width = 0.55
    y_max = max(plot_values) if plot_values else 0
    y_limit = y_max * 1.18 if y_max > 0 else 1
    depth_x = 0.10
    depth_y = y_limit * 0.025

    ax.set_xlim(-0.5, len(categories) - 0.5 + depth_x)
    ax.set_ylim(0, y_limit + depth_y)
    ax.set_xticks(range(len(categories)))
    ax.set_xticklabels(categories)

    bar_shapes = []
    for x_pos, height in enumerate(plot_values):
        left = x_pos - bar_width / 2
        right = x_pos + bar_width / 2

        front = Rectangle(
            (left, 0),
            bar_width,
            height,
            facecolor=color,
            edgecolor=shade_color(color, 0.72),
            linewidth=0.8,
            zorder=3,
        )
        right_side = Polygon(
            [(right, 0), (right + depth_x, depth_y), (right + depth_x, height + depth_y), (right, height)],
            closed=True,
            facecolor=shade_color(color, 0.72),
            edgecolor=shade_color(color, 0.62),
            linewidth=0.8,
            zorder=2.8,
        )
        top = Polygon(
            [(left, height), (left + depth_x, height + depth_y), (right + depth_x, height + depth_y), (right, height)],
            closed=True,
            facecolor=shade_color(color, 1.12),
            edgecolor=shade_color(color, 0.72),
            linewidth=0.8,
            zorder=3.2,
        )

        ax.add_patch(right_side)
        ax.add_patch(front)
        ax.add_patch(top)
        bar_shapes.append((x_pos, height))
    
    ax.grid(axis='y', color='#E0E0E0', linestyle='-', linewidth=0.8, zorder=0)
    ax.set_title(title, fontsize=18, fontweight='bold', pad=12, color='#2C3E50')
    ax.set_ylabel(ylabel, fontsize=14, color='#7F8C8D')
    ax.tick_params(axis='x', labelsize=14, colors='#2C3E50')
    ax.tick_params(axis='y', labelsize=14, colors='#7F8C8D')
    for label in ax.get_xticklabels():
        label.set_rotation(15)
        label.set_ha('right')
    
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#BDC3C7')
    ax.spines['bottom'].set_color('#BDC3C7')
    ax.set_axisbelow(True)
    
    for x_pos, height in bar_shapes:
        if height > 0:
            if is_time:
                total_minutes = int(round(height * 60))
                h = total_minutes // 60
                m = total_minutes % 60
                label = f"{h}ч {m}м" if h > 0 else f"{m}м"
            else:
                label = f"{int(height)}"
                
            ax.annotate(label,
                        xy=(x_pos + depth_x / 2, height + depth_y),
                        xytext=(0, 4),
                        textcoords="offset points",
                        ha='center', va='bottom', fontsize=8.5, fontweight='bold', color='#2C3E50')


def create_modern_chart(ws, anchor_cell: str, title: str, categories: List[str], values: List[float], color: str, is_time: bool = False):
    """
    Генерирует дизайнерский плоский график. 
    Размеры подобраны так, чтобы картинка занимала ровно 19 строк в Excel при стандартной высоте.
    """
    fig, ax = plt.subplots(figsize=(9.0, 5.8))
    draw_modern_chart(ax, title, categories, values, color, is_time)
    plt.tight_layout()
    
    img_buf = io.BytesIO()
    fig.savefig(img_buf, format='png', dpi=100) 
    plt.close(fig)
    img_buf.seek(0)
    
    xl_img = OpenpyxlImage(img_buf)
    ws.add_image(xl_img, anchor_cell)


def add_charts_pdf_page(pdf: PdfPages, page_title: str, categories: List[str], downtime_values: List[float], count_values: List[int]):
    fig, axes = plt.subplots(2, 1, figsize=(8.27, 11.69))
    fig.patch.set_facecolor('white')

    draw_modern_chart(
        axes[0],
        f"{page_title}: время остановок",
        categories,
        downtime_values,
        CHART_COLOR_DOWNTIME,
        is_time=True
    )
    draw_modern_chart(
        axes[1],
        f"{page_title}: количество остановок",
        categories,
        count_values,
        CHART_COLOR_COUNT,
        is_time=False
    )

    fig.subplots_adjust(left=0.15, right=0.9, top=0.9, bottom=0.15, hspace=0.55)
    pdf.savefig(fig)
    plt.close(fig)


def create_charts_pdf(output_path: Path, month: int, year: int, categories: List[str], monthly_downtime: List[float], monthly_stop_count: List[int], all_weeks_data: List[Tuple[int, List[float], List[int]]]):
    with PdfPages(output_path) as pdf:
        month_title = f"За {get_month_name_ru(month).lower()} {year}"
        add_charts_pdf_page(
            pdf,
            month_title,
            categories,
            monthly_downtime + [sum(monthly_downtime)],
            monthly_stop_count + [sum(monthly_stop_count)]
        )

        for week_num, w_downtime, w_stop_count in all_weeks_data:
            week_title = f"{week_num} неделя"
            add_charts_pdf_page(
                pdf,
                week_title,
                categories,
                w_downtime + [sum(w_downtime)],
                w_stop_count + [sum(w_stop_count)]
            )

def generate_report(month_str: str, config_path: str):
    config = load_config(Path(config_path))
    year, month = map(int, month_str.split("-"))
    
    stops_path_str = config.get("stops_input_path")
    if stops_path_str:
        workbook_path = Path(stops_path_str)
        if not workbook_path.is_absolute():
            workbook_path = Path(config_path).parent / workbook_path
    else:
        workbook_path = Path(config.get("output_workbook_path", config.get("workbook_path")))
        if not workbook_path.exists():
            workbook_path = Path(config.get("workbook_path"))
        
    wb_source = load_workbook(workbook_path, data_only=True)
    
    input_sheet_name = config.get("input_sheet_name", "Ввод_остановок")
    if input_sheet_name not in wb_source.sheetnames:
        print(f"Error: Sheet {input_sheet_name} not found.")
        return
        
    ws_input = wb_source[input_sheet_name]
    events_by_date_attr = {}
    
    for row in ws_input.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if not d: continue
        if isinstance(d, datetime.datetime):
            d = d.date()
        
        if d.year != year or d.month != month:
            continue
            
        attr = row[1]
        stop_t = parse_time(row[2])
        start_t = parse_time(row[3])
        reason = row[5]
        
        key = (d, attr)
        if key not in events_by_date_attr:
            events_by_date_attr[key] = []
        events_by_date_attr[key].append((stop_t, start_t, reason))

    wb = Workbook()
    sheet_name = f"{get_month_name_ru(month)}{year}"
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.showGridLines = False

    attractions = config.get("attractions", [])
    attr_names = [a["sheet_name"] for a in attractions]
    attr_full_names = [a.get("full_name", a["sheet_name"]) for a in attractions]
    attr_count = len(attr_names)
    
    for key in events_by_date_attr:
        events_by_date_attr[key].sort(key=lambda x: x[0] if x[0] else datetime.time(23, 59))
    
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(1, 1).value = "Дата"
    
    for i, name in enumerate(attr_full_names):
        start_col = 2 + i * 4
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 3)
        ws.cell(1, start_col).value = name
        
        ws.cell(2, start_col).value = "Время останов"
        ws.cell(2, start_col + 1).value = "Время старта"
        ws.cell(2, start_col + 2).value = "Время простоя"
        ws.cell(2, start_col + 3).value = "Причина остановки"

    summary_col = 2 + attr_count * 4
    ws.merge_cells(start_row=1, start_column=summary_col, end_row=2, end_column=summary_col)
    ws.cell(1, summary_col).value = "Всего время остановок"
    ws.merge_cells(start_row=1, start_column=summary_col + 1, end_row=2, end_column=summary_col + 1)
    ws.cell(1, summary_col + 1).value = "Всего Кол-во остановок"

    header_fill = PatternFill(start_color=RGB_HEADER_BLUE, end_color=RGB_HEADER_BLUE, fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for r in (1, 2):
        for c in range(1, summary_col + 2):
            cell = ws.cell(r, c)
            cell.fill = header_fill
            cell.alignment = center_align
            cell.font = Font(bold=True, size=10)
            apply_border(cell)

    current_row = 3
    first_day = datetime.date(year, month, 1)
    if month == 12:
        last_day = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
    else:
        last_day = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
    
    monthly_downtime = [0.0] * attr_count
    monthly_stop_count = [0] * attr_count
    
    weekly_downtime = [0.0] * attr_count
    weekly_stop_count = [0] * attr_count
    all_weeks_data = []

    for d in (first_day + datetime.timedelta(n) for n in range((last_day - first_day).days + 1)):
        # Determine park hours for this day
        if d.weekday() < 4: # Mon-Thu
            po_str = config.get("park_open_mon_thu", "09:00")
            pc_str = config.get("park_close_mon_thu", "22:00")
        else: # Fri-Sun
            po_str = config.get("park_open_fri_sun", "09:00")
            pc_str = config.get("park_close_fri_sun", "23:00")
        
        day_park_open = parse_time(po_str)
        day_park_close = parse_time(pc_str)

        day_events = []
        max_ev = 1
        for attr in attr_names:
            evs = events_by_date_attr.get((d, attr), [])
            day_events.append(evs)
            max_ev = max(max_ev, len(evs))
            
        ws.cell(current_row, 1).value = d
        orange_fill = PatternFill(start_color=RGB_ORANGE, end_color=RGB_ORANGE, fill_type="solid")
        for i in range(attr_count):
            start_col = 2 + i * 4
            ws.cell(current_row, start_col + 1).value = day_park_open
            ws.cell(current_row, start_col + 1).number_format = "h:mm"
            for c_off in range(3):
                cell = ws.cell(current_row, start_col + c_off)
                cell.fill = orange_fill
                
        for c in range(1, summary_col + 2):
            cell = ws.cell(current_row, c)
            apply_border(cell)
            cell.alignment = center_align
        ws.cell(current_row, 1).number_format = "dd.mm.yy"
        current_row += 1
        
        day_dt_sums = [0.0] * attr_count
        day_sc_sums = [0] * attr_count
        
        for ev_idx in range(max_ev):
            ws.cell(current_row, 1).value = d
            ws.cell(current_row, 1).number_format = "dd.mm.yy"
            for i in range(attr_count):
                start_col = 2 + i * 4
                if ev_idx < len(day_events[i]):
                    stop_t, start_t, reason = day_events[i][ev_idx]
                    ws.cell(current_row, start_col).value = stop_t
                    ws.cell(current_row, start_col + 1).value = start_t
                    ws.cell(current_row, start_col + 3).value = reason
                    
                    if stop_t and start_t:
                        dt_val = calculate_downtime(stop_t, start_t)
                        ws.cell(current_row, start_col + 2).value = dt_val
                        day_dt_sums[i] += dt_val
                        day_sc_sums[i] += 1
                    
                    ws.cell(current_row, start_col).number_format = "h:mm"
                    ws.cell(current_row, start_col + 1).number_format = "h:mm"
                    ws.cell(current_row, start_col + 2).number_format = "h:mm"
            
            for c in range(1, summary_col + 2):
                cell = ws.cell(current_row, c)
                apply_border(cell)
                cell.alignment = center_align
            current_row += 1
            
        ws.cell(current_row, 1).value = d
        ws.cell(current_row, 1).number_format = "dd.mm.yy"
        for i in range(attr_count):
            start_col = 2 + i * 4
            ws.cell(current_row, start_col).value = day_park_close
            ws.cell(current_row, start_col).number_format = "h:mm"
            for c_off in range(3):
                cell = ws.cell(current_row, start_col + c_off)
                cell.fill = orange_fill
        
        for c in range(1, summary_col + 2):
            cell = ws.cell(current_row, c)
            apply_border(cell)
            cell.alignment = center_align
        current_row += 1
        
        ws.cell(current_row, 1).value = d
        ws.cell(current_row, 1).number_format = "dd.mm.yy"
        yellow_fill = PatternFill(start_color=RGB_YELLOW, end_color=RGB_YELLOW, fill_type="solid")
        
        day_total_dt = sum(day_dt_sums)
        day_total_sc = sum(day_sc_sums)
        
        for i in range(attr_count):
            start_col = 2 + i * 4
            ws.cell(current_row, start_col + 2).value = day_dt_sums[i]
            ws.cell(current_row, start_col + 3).value = day_sc_sums[i]
            ws.cell(current_row, start_col + 2).number_format = "h:mm"
            
            monthly_downtime[i] += day_dt_sums[i]
            monthly_stop_count[i] += day_sc_sums[i]
            weekly_downtime[i] += day_dt_sums[i]
            weekly_stop_count[i] += day_sc_sums[i]
            
        ws.cell(current_row, summary_col).value = day_total_dt
        ws.cell(current_row, summary_col).number_format = "[h]:mm:ss"
        ws.cell(current_row, summary_col + 1).value = day_total_sc
        
        for c in range(1, summary_col + 2):
            cell = ws.cell(current_row, c)
            cell.fill = yellow_fill
            apply_border(cell)
            cell.alignment = center_align
        current_row += 1
        
        if d.weekday() == 6 or d == last_day:
            week_num = d.isocalendar()[1]
            all_weeks_data.append((week_num, list(weekly_downtime), list(weekly_stop_count)))
            
            ws.cell(current_row, 1).value = f"{week_num} неделя"
            green_fill = PatternFill(start_color=RGB_GREEN, end_color=RGB_GREEN, fill_type="solid")
            
            for i in range(attr_count):
                start_col = 2 + i * 4
                ws.cell(current_row, start_col + 2).value = weekly_downtime[i]
                ws.cell(current_row, start_col + 3).value = weekly_stop_count[i]
                ws.cell(current_row, start_col + 2).number_format = "h:mm"
                
                weekly_downtime[i] = 0.0
                weekly_stop_count[i] = 0
            
            ws.cell(current_row, summary_col).value = sum(ws.cell(current_row, 2 + i * 4 + 2).value for i in range(attr_count))
            ws.cell(current_row, summary_col).number_format = "[h]:mm:ss"
            ws.cell(current_row, summary_col + 1).value = sum(ws.cell(current_row, 2 + i * 4 + 3).value for i in range(attr_count))
            
            for c in range(1, summary_col + 2):
                cell = ws.cell(current_row, c)
                cell.fill = green_fill
                apply_border(cell)
                cell.alignment = center_align
            current_row += 1

    # Строка Итого за месяц
    ws.cell(current_row, 1).value = f"За {get_month_name_ru(month).lower()}"
    blue_fill = PatternFill(start_color=RGB_BLUE, end_color=RGB_BLUE, fill_type="solid")
    for i in range(attr_count):
        start_col = 2 + i * 4
        ws.cell(current_row, start_col + 2).value = monthly_downtime[i]
        ws.cell(current_row, start_col + 3).value = monthly_stop_count[i]
        ws.cell(current_row, start_col + 2).number_format = "h:mm"
        
    ws.cell(current_row, summary_col).value = sum(monthly_downtime)
    ws.cell(current_row, summary_col).number_format = "[h]:mm:ss"
    ws.cell(current_row, summary_col + 1).value = sum(monthly_stop_count)
    
    for c in range(1, summary_col + 2):
        cell = ws.cell(current_row, c)
        cell.fill = blue_fill
        apply_border(cell)
        cell.alignment = center_align
        cell.font = Font(bold=True)

    ws.freeze_panes = "B3"
    
    for c in range(1, summary_col + 2):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.column_dimensions['A'].width = 12

    chart_categories = attr_names + ["Итого"]

    # Сохранение итогового отчета
    report_dir = Path(config.get("report_dir", "reports"))
    report_dir.mkdir(exist_ok=True)
    
    output_filename = f"Отчет_{sheet_name}.xlsx"
    output_path = report_dir / output_filename
    charts_output_path = report_dir / f"Графики_{sheet_name}.pdf"
    
    attempts = 0
    while attempts < 10:
        try:
            wb.save(output_path)
            create_charts_pdf(
                charts_output_path,
                month,
                year,
                chart_categories,
                monthly_downtime,
                monthly_stop_count,
                all_weeks_data
            )
            print(f"Table report generated: {output_path}")
            print(f"Charts PDF generated: {charts_output_path}")
            break
        except PermissionError:
            print(f"Error: Could not save to {output_path}. Retrying...")
            output_path = output_path.with_name(f"REPORT_{output_path.name}")
            attempts += 1

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate clean monthly attraction report with side-by-side charts.")
    parser.add_argument("--month", required=True, help="Month to process (YYYY-MM)")
    parser.add_argument("--config", default="config.json", help="Path to config.json")
    args = parser.parse_args()
    
    generate_report(args.month, args.config)
