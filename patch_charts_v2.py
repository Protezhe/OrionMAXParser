import sys

def patch():
    with open('generate_monthly_report.py', 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    new_lines = []
    skip_next = False
    for i in range(len(lines)):
        if skip_next:
            skip_next = False
            continue
        # Remove duplicate import
        if "from openpyxl.chart.label import DataLabelList" in lines[i]:
            if i + 1 < len(lines) and "from openpyxl.chart.label import DataLabelList" in lines[i+1]:
                new_lines.append(lines[i])
                skip_next = True
                continue
        new_lines.append(lines[i])
    
    content = "".join(new_lines)
    
    # Define the correctly styled blocks for all 4 types of charts
    
    # 1. Monthly Downtime
    m1_new = """    # 1. Monthly Downtime Chart (3D)
    dt_chart = BarChart3D()
    dt_chart.type = "col"
    dt_chart.grouping = "clustered"
    dt_chart.title = f"Время остановок (ч) - {get_month_name_ru(month)} {year}"
    dt_chart.title.layout = Layout(manualLayout=ManualLayout(xMode='edge', yMode='edge', x=0.4402, y=0.0))
    dt_chart.legend = None
    dt_chart.width = 15
    dt_chart.height = 7.5
    
    dt_data = Reference(ws, min_col=2, min_row=summary_header_row + 1, max_row=summary_end_row)
    dt_cats = Reference(ws, min_col=4, min_row=summary_header_row + 1, max_row=summary_end_row) # Use combined labels
    dt_chart.add_data(dt_data, titles_from_data=False)
    dt_chart.set_categories(dt_cats)
    
    dt_chart.x_axis.delete = False
    dt_chart.y_axis.delete = False
    dt_chart.dLbls = DataLabelList()
    dt_chart.dLbls.showCatName = True
    
    if dt_chart.series:
        from openpyxl.drawing.colors import ColorChoice
        s = dt_chart.series[0]
        s.graphicalProperties = GraphicalProperties(solidFill=ColorChoice(srgbClr="FFC000"))

    ws.add_chart(dt_chart, f"{get_column_letter(summary_col + 2)}4")"""

    # 2. Monthly Stop Count
    m2_new = """    # 2. Monthly Stop Count Chart (3D)
    sc_chart = BarChart3D()
    sc_chart.type = "col"
    sc_chart.grouping = "clustered"
    sc_chart.title = f"Количество остановок - {get_month_name_ru(month)} {year}"
    sc_chart.title.layout = Layout(manualLayout=ManualLayout(xMode='edge', yMode='edge', x=0.3673, y=0.0))
    sc_chart.legend = None
    sc_chart.width = 15
    sc_chart.height = 7.5
    
    sc_data = Reference(ws, min_col=3, min_row=summary_header_row + 1, max_row=summary_end_row)
    sc_cats = Reference(ws, min_col=5, min_row=summary_header_row + 1, max_row=summary_end_row) # Use combined labels
    sc_chart.add_data(sc_data, titles_from_data=False)
    sc_chart.set_categories(sc_cats)
    
    sc_chart.x_axis.delete = False
    sc_chart.y_axis.delete = False
    sc_chart.dLbls = DataLabelList()
    sc_chart.dLbls.showCatName = True
    
    if sc_chart.series:
        from openpyxl.drawing.colors import ColorChoice
        s = sc_chart.series[0]
        s.graphicalProperties = GraphicalProperties(solidFill=ColorChoice(srgbClr="4F81BD"))

    ws.add_chart(sc_chart, f"{get_column_letter(summary_col + 2)}15")"""

    # 3. Weekly Downtime
    w1_new = """        # Weekly Downtime Chart
        w_dt_chart = BarChart3D()
        w_dt_chart.type = "col"
        w_dt_chart.grouping = "clustered"
        w_dt_chart.title = f"Время остановок (ч) - {week_num} неделю"
        w_dt_chart.legend = None
        w_dt_chart.width = 15
        w_dt_chart.height = 7.5
        
        w_dt_data = Reference(ws, min_col=2, min_row=w_header_row + 1, max_row=w_end_row)
        w_dt_cats = Reference(ws, min_col=4, min_row=w_header_row + 1, max_row=w_end_row)
        w_dt_chart.add_data(w_dt_data, titles_from_data=False)
        w_dt_chart.set_categories(w_dt_cats)
        
        w_dt_chart.x_axis.delete = False
        w_dt_chart.y_axis.delete = False
        w_dt_chart.dLbls = DataLabelList()
        w_dt_chart.dLbls.showCatName = True
        
        if w_dt_chart.series:
            from openpyxl.drawing.colors import ColorChoice
            s = w_dt_chart.series[0]
            s.graphicalProperties = GraphicalProperties(solidFill=ColorChoice(srgbClr="FFC000"))

        chart_row_offset = 26 + idx * 22
        ws.add_chart(w_dt_chart, f"{get_column_letter(summary_col + 2)}{chart_row_offset}")"""

    # 4. Weekly Stop Count
    w2_new = """        # Weekly Stop Count Chart
        w_sc_chart = BarChart3D()
        w_sc_chart.type = "col"
        w_sc_chart.grouping = "clustered"
        w_sc_chart.title = f"Количество остановок - {week_num} неделю"
        w_sc_chart.legend = None
        w_sc_chart.width = 15
        w_sc_chart.height = 7.5
        
        w_sc_data = Reference(ws, min_col=3, min_row=w_header_row + 1, max_row=w_end_row)
        w_sc_cats = Reference(ws, min_col=5, min_row=w_header_row + 1, max_row=w_end_row)
        w_sc_chart.add_data(w_sc_data, titles_from_data=False)
        w_sc_chart.set_categories(w_sc_cats)
        
        w_sc_chart.x_axis.delete = False
        w_sc_chart.y_axis.delete = False
        w_sc_chart.dLbls = DataLabelList()
        w_sc_chart.dLbls.showCatName = True
        
        if w_sc_chart.series:
            from openpyxl.drawing.colors import ColorChoice
            s = w_sc_chart.series[0]
            s.graphicalProperties = GraphicalProperties(solidFill=ColorChoice(srgbClr="4F81BD"))

        ws.add_chart(w_sc_chart, f"{get_column_letter(summary_col + 2)}{chart_row_offset + 11}")"""

    # We need to find the old blocks carefully. Since they might have been partially patched or vary, 
    # we'll use a more robust search.
    
    import re
    
    # Match blocks starting with comments and ending with ws.add_chart
    # Monthly Downtime
    m1_pattern = re.compile(r'    # 1\. Monthly Downtime Chart \(3D\).*?ws\.add_chart\(dt_chart, f"{get_column_letter\(summary_col \+ 2\)}4"\)', re.DOTALL)
    content = m1_pattern.sub(m1_new, content)
    
    # Monthly Stop Count
    m2_pattern = re.compile(r'    # 2\. Monthly Stop Count Chart \(3D\).*?ws\.add_chart\(sc_chart, f"{get_column_letter\(summary_col \+ 2\)}15"\)', re.DOTALL)
    content = m2_pattern.sub(m2_new, content)
    
    # Weekly Downtime
    w1_pattern = re.compile(r'        # Weekly Downtime Chart.*?ws\.add_chart\(w_dt_chart, f"{get_column_letter\(summary_col \+ 2\)}{chart_row_offset}"\)', re.DOTALL)
    content = w1_pattern.sub(w1_new, content)
    
    # Weekly Stop Count
    w2_pattern = re.compile(r'        # Weekly Stop Count Chart.*?ws\.add_chart\(w_sc_chart, f"{get_column_letter\(summary_col \+ 2\)}{chart_row_offset \+ 11}"\)', re.DOTALL)
    content = w2_pattern.sub(w2_new, content)

    with open('generate_monthly_report.py', 'w', encoding='utf-8') as f:
        f.write(content)

if __name__ == '__main__':
    patch()
