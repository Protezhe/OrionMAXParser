import argparse
import datetime
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart3D, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice
# Removed incorrect import
from openpyxl.styles import Alignment, Border, Color, Fill, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.chart.data_source import AxDataSource, StrRef

# Replicating VBA constants and colors
RGB_ORANGE = "FFC000"
RGB_YELLOW = "FFFF00"
RGB_GREEN = "70AD47"
RGB_BLUE = "00B0F0"
RGB_HEADER_BLUE = "D9EAF7"

def load_config(config_path: Path) -> Dict[str, Any]:
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)

def get_month_name_ru(month: int) -> str:
    months = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    ]
    return months[month - 1]

def parse_time(value: Any) -> datetime.time | None:
    if isinstance(value, datetime.time):
        return value
    if isinstance(value, datetime.datetime):
        return value.time()
    if isinstance(value, str):
        try:
            return datetime.datetime.strptime(value, "%H:%M").time()
        except ValueError:
            return None
    return None

def time_to_seconds(t: datetime.time) -> int:
    return t.hour * 3600 + t.minute * 60 + t.second

def seconds_to_excel_time(seconds: float) -> float:
    # Excel time is a fraction of a day
    return seconds / (24 * 3600)

def calculate_downtime(start: datetime.time, end: datetime.time) -> float:
    s = time_to_seconds(start)
    e = time_to_seconds(end)
    if e < s:
        # Crosses midnight
        diff = (24 * 3600 - s) + e
    else:
        diff = e - s
    return seconds_to_excel_time(diff)

def apply_border(cell):
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def generate_report(month_str: str, config_path: str):
    config = load_config(Path(config_path))
    year, month = map(int, month_str.split("-"))
    
    stops_path_str = config.get("stops_input_path")
    if stops_path_str:
        workbook_path = Path(stops_path_str)
        if not workbook_path.is_absolute():
            workbook_path = Path(config_path).parent / workbook_path
    else:
        workbook_path = Path(config.get("output_workbook_path", config.get("workbook_path")))
        if not workbook_path.exists():
            workbook_path = Path(config.get("workbook_path"))
        
    # Load source data only
    wb_source = load_workbook(workbook_path, data_only=True)
    
    input_sheet_name = config.get("input_sheet_name", "Ввод_остановок")
    if input_sheet_name not in wb_source.sheetnames:
        print(f"Error: Sheet {input_sheet_name} not found.")
        return
        
    ws_input = wb_source[input_sheet_name]
    
    # Load data from input sheet
    events_by_date_attr = {} # (date, attraction) -> list of (stop, start, reason)
    
    for row in ws_input.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if not d: continue
        if isinstance(d, datetime.datetime):
            d = d.date()
        
        if d.year != year or d.month != month:
            continue
            
        attr = row[1]
        stop_t = parse_time(row[2])
        start_t = parse_time(row[3])
        reason = row[5]
        
        key = (d, attr)
        if key not in events_by_date_attr:
            events_by_date_attr[key] = []
        events_by_date_attr[key].append((stop_t, start_t, reason))

    # Create NEW workbook for report
    wb = Workbook()
    sheet_name = f"{get_month_name_ru(month)}{year}"
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.showGridLines = False

    attractions = config.get("attractions", [])
    attr_names = [a["sheet_name"] for a in attractions]
    attr_full_names = [a.get("full_name", a["sheet_name"]) for a in attractions]
    attr_count = len(attr_names)
    
    # Sort events by stop time
    for key in events_by_date_attr:
        events_by_date_attr[key].sort(key=lambda x: x[0] if x[0] else datetime.time(23, 59))
    
    # Header Construction
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(1, 1).value = "Дата"
    
    for i, name in enumerate(attr_full_names):
        start_col = 2 + i * 4
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 3)
        ws.cell(1, start_col).value = name
        
        ws.cell(2, start_col).value = "Время останов"
        ws.cell(2, start_col + 1).value = "Время старта"
        ws.cell(2, start_col + 2).value = "Время простоя"
        ws.cell(2, start_col + 3).value = "Причина остановки"

    summary_col = 2 + attr_count * 4
    ws.merge_cells(start_row=1, start_column=summary_col, end_row=2, end_column=summary_col)
    ws.cell(1, summary_col).value = "Всего время остановок"
    ws.merge_cells(start_row=1, start_column=summary_col + 1, end_row=2, end_column=summary_col + 1)
    ws.cell(1, summary_col + 1).value = "Всего Кол-во остановок"

    # Styling header
    header_fill = PatternFill(start_color=RGB_HEADER_BLUE, end_color=RGB_HEADER_BLUE, fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for r in (1, 2):
        for c in range(1, summary_col + 2):
            cell = ws.cell(r, c)
            cell.fill = header_fill
            cell.alignment = center_align
            cell.font = Font(bold=True, size=10)
            apply_border(cell)

    # Data Rows
    current_row = 3
    first_day = datetime.date(year, month, 1)
    if month == 12:
        last_day = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
    else:
        last_day = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
    
    monthly_downtime = [0.0] * attr_count
    monthly_stop_count = [0] * attr_count
    
    weekly_downtime = [0.0] * attr_count
    weekly_stop_count = [0] * attr_count
    all_weeks_data = [] # List of (week_num, downtime_list, stop_count_list)

    for d in (first_day + datetime.timedelta(n) for n in range((last_day - first_day).days + 1)):
        # Determine park hours for this day
        if d.weekday() < 4: # Mon-Thu
            po_str = config.get("park_open_mon_thu", "09:00")
            pc_str = config.get("park_close_mon_thu", "22:00")
        else: # Fri-Sun
            po_str = config.get("park_open_fri_sun", "09:00")
            pc_str = config.get("park_close_fri_sun", "23:00")
        
        day_park_open = parse_time(po_str)
        day_park_close = parse_time(pc_str)

        # Determine max events for this day
        day_events = []
        max_ev = 1
        for attr in attr_names:
            evs = events_by_date_attr.get((d, attr), [])
            day_events.append(evs)
            max_ev = max(max_ev, len(evs))
            
        # 1. Start of day row (Orange)
        ws.cell(current_row, 1).value = d
        orange_fill = PatternFill(start_color=RGB_ORANGE, end_color=RGB_ORANGE, fill_type="solid")
        for i in range(attr_count):
            start_col = 2 + i * 4
            ws.cell(current_row, start_col + 1).value = day_park_open
            ws.cell(current_row, start_col + 1).number_format = "h:mm"
            # Fill color for stop/start/downtime columns as in VBA
            for c_off in range(3):
                cell = ws.cell(current_row, start_col + c_off)
                cell.fill = orange_fill
                
        for c in range(1, summary_col + 2):
            cell = ws.cell(current_row, c)
            apply_border(cell)
            cell.alignment = center_align
        ws.cell(current_row, 1).number_format = "dd.mm.yy"
        current_row += 1
        
        # 2. Event rows
        day_dt_sums = [0.0] * attr_count
        day_sc_sums = [0] * attr_count
        
        for ev_idx in range(max_ev):
            ws.cell(current_row, 1).value = d
            ws.cell(current_row, 1).number_format = "dd.mm.yy"
            for i in range(attr_count):
                start_col = 2 + i * 4
                if ev_idx < len(day_events[i]):
                    stop_t, start_t, reason = day_events[i][ev_idx]
                    ws.cell(current_row, start_col).value = stop_t
                    ws.cell(current_row, start_col + 1).value = start_t
                    ws.cell(current_row, start_col + 3).value = reason
                    
                    if stop_t and start_t:
                        dt_val = calculate_downtime(stop_t, start_t)
                        ws.cell(current_row, start_col + 2).value = dt_val
                        day_dt_sums[i] += dt_val
                        day_sc_sums[i] += 1
                    
                    ws.cell(current_row, start_col).number_format = "h:mm"
                    ws.cell(current_row, start_col + 1).number_format = "h:mm"
                    ws.cell(current_row, start_col + 2).number_format = "h:mm"
            
            for c in range(1, summary_col + 2):
                cell = ws.cell(current_row, c)
                apply_border(cell)
                cell.alignment = center_align
            current_row += 1
            
        # 3. End of day row (Orange)
        ws.cell(current_row, 1).value = d
        ws.cell(current_row, 1).number_format = "dd.mm.yy"
        for i in range(attr_count):
            start_col = 2 + i * 4
            ws.cell(current_row, start_col).value = day_park_close
            ws.cell(current_row, start_col).number_format = "h:mm"
            for c_off in range(3):
                cell = ws.cell(current_row, start_col + c_off)
                cell.fill = orange_fill
        
        for c in range(1, summary_col + 2):
            cell = ws.cell(current_row, c)
            apply_border(cell)
            cell.alignment = center_align
        current_row += 1
        
        # 4. Day Total Row (Yellow)
        ws.cell(current_row, 1).value = d
        ws.cell(current_row, 1).number_format = "dd.mm.yy"
        yellow_fill = PatternFill(start_color=RGB_YELLOW, end_color=RGB_YELLOW, fill_type="solid")
        
        day_total_dt = sum(day_dt_sums)
        day_total_sc = sum(day_sc_sums)
        
        for i in range(attr_count):
            start_col = 2 + i * 4
            # VBA fills normative time here? No, it fills total downtime for the attraction for the day
            ws.cell(current_row, start_col + 2).value = day_dt_sums[i]
            ws.cell(current_row, start_col + 3).value = day_sc_sums[i]
            ws.cell(current_row, start_col + 2).number_format = "h:mm"
            
            monthly_downtime[i] += day_dt_sums[i]
            monthly_stop_count[i] += day_sc_sums[i]
            weekly_downtime[i] += day_dt_sums[i]
            weekly_stop_count[i] += day_sc_sums[i]
            
        ws.cell(current_row, summary_col).value = day_total_dt
        ws.cell(current_row, summary_col).number_format = "[h]:mm:ss"
        ws.cell(current_row, summary_col + 1).value = day_total_sc
        
        for c in range(1, summary_col + 2):
            cell = ws.cell(current_row, c)
            cell.fill = yellow_fill
            apply_border(cell)
            cell.alignment = center_align
        current_row += 1
        
        # 5. Weekly Total Row (Green)
        # If it's Sunday or last day of month
        if d.weekday() == 6 or d == last_day:
            week_num = d.isocalendar()[1]
            all_weeks_data.append((week_num, list(weekly_downtime), list(weekly_stop_count)))
            
            ws.cell(current_row, 1).value = f"{week_num} неделя"
            green_fill = PatternFill(start_color=RGB_GREEN, end_color=RGB_GREEN, fill_type="solid")
            
            for i in range(attr_count):
                start_col = 2 + i * 4
                ws.cell(current_row, start_col + 2).value = weekly_downtime[i]
                ws.cell(current_row, start_col + 3).value = weekly_stop_count[i]
                ws.cell(current_row, start_col + 2).number_format = "h:mm"
                
                weekly_downtime[i] = 0.0
                weekly_stop_count[i] = 0
            
            ws.cell(current_row, summary_col).value = sum(ws.cell(current_row, 2 + i * 4 + 2).value for i in range(attr_count))
            ws.cell(current_row, summary_col).number_format = "[h]:mm:ss"
            ws.cell(current_row, summary_col + 1).value = sum(ws.cell(current_row, 2 + i * 4 + 3).value for i in range(attr_count))
            
            for c in range(1, summary_col + 2):
                cell = ws.cell(current_row, c)
                cell.fill = green_fill
                apply_border(cell)
                cell.alignment = center_align
            current_row += 1

    # Grand Total Row (Blue)
    ws.cell(current_row, 1).value = f"За {get_month_name_ru(month).lower()}"
    blue_fill = PatternFill(start_color=RGB_BLUE, end_color=RGB_BLUE, fill_type="solid")
    for i in range(attr_count):
        start_col = 2 + i * 4
        ws.cell(current_row, start_col + 2).value = monthly_downtime[i]
        ws.cell(current_row, start_col + 3).value = monthly_stop_count[i]
        ws.cell(current_row, start_col + 2).number_format = "h:mm"
        
    ws.cell(current_row, summary_col).value = sum(monthly_downtime)
    ws.cell(current_row, summary_col).number_format = "[h]:mm:ss"
    ws.cell(current_row, summary_col + 1).value = sum(monthly_stop_count)
    
    for c in range(1, summary_col + 2):
        cell = ws.cell(current_row, c)
        cell.fill = blue_fill
        apply_border(cell)
        cell.alignment = center_align
        cell.font = Font(bold=True)
    current_row += 1

    # Freeze panes
    ws.freeze_panes = "B3"
    
    # Auto-width
    for c in range(1, summary_col + 2):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.column_dimensions['A'].width = 12

    # --- Charts (Monthly Summary per Attraction) ---
    # In 2025, charts show totals for each attraction for the whole month.
    # We will create a small table for this.
    
    summary_data_row = current_row + 2
    ws.cell(summary_data_row, 1).value = "Итоги по аттракционам"
    ws.cell(summary_data_row, 1).font = Font(bold=True)
    
    summary_header_row = summary_data_row + 1
    ws.cell(summary_header_row, 1).value = "Аттракцион"
    ws.cell(summary_header_row, 2).value = "Простой (ч)"
    ws.cell(summary_header_row, 3).value = "Кол-во ост."
    ws.cell(summary_header_row, 4).value = "Подпись для графика (простой)"
    ws.cell(summary_header_row, 5).value = "Подпись для графика (кол-во)"
    ws.cell(summary_header_row, 6).value = "Название для оси"
    ws.cell(summary_header_row, 7).value = "Простой для графика"
    
    def format_duration_ru(dt_excel):
        total_seconds = int(round(dt_excel * 24 * 3600))
        h = total_seconds // 3600
        m = (total_seconds % 3600) // 60
        return f"{h} ч. {m} мин."

    chart_left_col = get_column_letter(summary_col + 2)
    chart_height = 7.5
    chart_width = 15
    chart_row_span = 21
    chart_pair_row_span = 50

    def add_summary_chart(
        anchor_row: int,
        title: str,
        data_col: int,
        category_col: int,
        min_row: int,
        max_row: int,
        color: str,
        number_format: str,
    ) -> None:
        chart = BarChart3D()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = title
        chart.legend = None
        chart.width = chart_width
        chart.height = chart_height
        chart.gapWidth = 150
        chart.gapDepth = 150
        chart.y_axis.numFmt = number_format

        data = Reference(ws, min_col=data_col, min_row=min_row, max_row=max_row)
        category_col_letter = get_column_letter(category_col)
        category_ref = (
            f"{quote_sheetname(ws.title)}!"
            f"${category_col_letter}${min_row}:${category_col_letter}${max_row}"
        )
        chart.add_data(data, titles_from_data=False)

        chart.x_axis.delete = False
        chart.x_axis.axPos = "b"
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.lblOffset = 100
        chart.y_axis.delete = False
        chart.y_axis.axPos = "l"
        chart.y_axis.tickLblPos = "nextTo"

        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showCatName = False
        chart.dLbls.showSerName = False
        chart.dLbls.showLegendKey = False
        chart.dLbls.showPercent = False
        chart.dLbls.showBubbleSize = False
        chart.dLbls.showLeaderLines = False
        chart.dLbls.numFmt = number_format

        if chart.series:
            series = chart.series[0]
            series.tx = SeriesLabel(v=title)
            series.dLbls = DataLabelList()
            series.dLbls.showVal = True
            series.dLbls.showCatName = False
            series.dLbls.showSerName = False
            series.dLbls.showLegendKey = False
            series.dLbls.showPercent = False
            series.dLbls.showBubbleSize = False
            series.dLbls.showLeaderLines = False
            series.dLbls.numFmt = number_format
            series.cat = AxDataSource(strRef=StrRef(f=category_ref))
            series.graphicalProperties = GraphicalProperties(solidFill=ColorChoice(srgbClr=color))

        ws.add_chart(chart, f"{chart_left_col}{anchor_row}")

    for i in range(attr_count):
        row = summary_header_row + 1 + i
        ws.cell(row, 1).value = attr_full_names[i]
        dt_h = monthly_downtime[i] * 24
        sc = monthly_stop_count[i]
        ws.cell(row, 2).value = dt_h
        ws.cell(row, 3).value = sc
        ws.cell(row, 2).number_format = "0.00"
        
        # Format labels: "Name; H ч. M мин."
        ws.cell(row, 4).value = f"{attr_full_names[i]}; {format_duration_ru(monthly_downtime[i])}"
        ws.cell(row, 5).value = f"{attr_full_names[i]}; {sc}"
        ws.cell(row, 6).value = attr_names[i]
        ws.cell(row, 7).value = monthly_downtime[i]
        ws.cell(row, 7).number_format = "[h]:mm"

    # Add Grand Total to summary table
    summary_end_row = summary_header_row + attr_count + 1
    ws.cell(summary_end_row, 1).value = "Итого"
    total_dt = sum(monthly_downtime)
    total_sc = sum(monthly_stop_count)
    ws.cell(summary_end_row, 2).value = total_dt * 24
    ws.cell(summary_end_row, 3).value = total_sc
    ws.cell(summary_end_row, 2).number_format = "0.00"
    ws.cell(summary_end_row, 4).value = f"Итого; {format_duration_ru(total_dt)}"
    ws.cell(summary_end_row, 5).value = f"Итого; {total_sc}"
    ws.cell(summary_end_row, 6).value = "Итого"
    ws.cell(summary_end_row, 7).value = total_dt
    ws.cell(summary_end_row, 7).number_format = "[h]:mm"
    
    # 1. Monthly Downtime Chart
    add_summary_chart(
        anchor_row=4,
        title=f"Время остановок эксплуатации аттракционов за {get_month_name_ru(month).lower()} {year}",
        data_col=7,
        category_col=6,
        min_row=summary_header_row + 1,
        max_row=summary_end_row,
        color="FFC000",
        number_format="[h]:mm",
    )

    # 2. Monthly Stop Count Chart
    add_summary_chart(
        anchor_row=4 + chart_row_span,
        title=f"Количество остановок эксплуатации аттракционов за {get_month_name_ru(month).lower()} {year}",
        data_col=3,
        category_col=6,
        min_row=summary_header_row + 1,
        max_row=summary_end_row,
        color="4F81BD",
        number_format="0",
    )

    # --- Weekly Charts ---
    running_row = summary_end_row + 4
    for idx, (week_num, w_downtime, w_stop_count) in enumerate(all_weeks_data):
        ws.cell(running_row, 1).value = f"Итоги за {week_num} неделю"
        ws.cell(running_row, 1).font = Font(bold=True)
        
        w_header_row = running_row + 1
        ws.cell(w_header_row, 1).value = "Аттракцион"
        ws.cell(w_header_row, 2).value = "Простой (ч)"
        ws.cell(w_header_row, 3).value = "Кол-во ост."
        ws.cell(w_header_row, 4).value = "Подпись для графика (простой)"
        ws.cell(w_header_row, 5).value = "Подпись для графика (кол-во)"
        ws.cell(w_header_row, 6).value = "Название для оси"
        ws.cell(w_header_row, 7).value = "Простой для графика"
        
        for i in range(attr_count):
            row = w_header_row + 1 + i
            ws.cell(row, 1).value = attr_full_names[i]
            dt_h = w_downtime[i] * 24
            sc = w_stop_count[i]
            ws.cell(row, 2).value = dt_h
            ws.cell(row, 3).value = sc
            ws.cell(row, 2).number_format = "0.00"
            ws.cell(row, 4).value = f"{attr_full_names[i]}; {format_duration_ru(w_downtime[i])}"
            ws.cell(row, 5).value = f"{attr_full_names[i]}; {sc}"
            ws.cell(row, 6).value = attr_names[i]
            ws.cell(row, 7).value = w_downtime[i]
            ws.cell(row, 7).number_format = "[h]:mm"
            
        # Weekly Grand Total
        w_end_row = w_header_row + attr_count + 1
        ws.cell(w_end_row, 1).value = "Итого"
        w_total_dt = sum(w_downtime)
        w_total_sc = sum(w_stop_count)
        ws.cell(w_end_row, 2).value = w_total_dt * 24
        ws.cell(w_end_row, 3).value = w_total_sc
        ws.cell(w_end_row, 2).number_format = "0.00"
        w_end_row_4_val = f"Итого; {format_duration_ru(w_total_dt)}"
        ws.cell(w_end_row, 4).value = w_end_row_4_val
        w_end_row_5_val = f"Итого; {w_total_sc}"
        ws.cell(w_end_row, 5).value = w_end_row_5_val
        ws.cell(w_end_row, 6).value = "Итого"
        ws.cell(w_end_row, 7).value = w_total_dt
        ws.cell(w_end_row, 7).number_format = "[h]:mm"
        
        chart_row_offset = 4 + chart_pair_row_span + idx * chart_pair_row_span

        # Weekly Downtime Chart
        add_summary_chart(
            anchor_row=chart_row_offset,
            title=f"Время остановок эксплуатации аттракционов за {week_num} неделю",
            data_col=7,
            category_col=6,
            min_row=w_header_row + 1,
            max_row=w_end_row,
            color="FFC000",
            number_format="[h]:mm",
        )

        # Weekly Stop Count Chart
        add_summary_chart(
            anchor_row=chart_row_offset + chart_row_span,
            title=f"Количество остановок эксплуатации аттракционов за {week_num} неделю",
            data_col=3,
            category_col=6,
            min_row=w_header_row + 1,
            max_row=w_end_row,
            color="4F81BD",
            number_format="0",
        )
        
        running_row = w_end_row + 2

    # Save as .xlsx
    report_dir = Path(config.get("report_dir", "reports"))
    report_dir.mkdir(exist_ok=True)
    
    output_filename = f"Отчет_{sheet_name}.xlsx"
    output_path = report_dir / output_filename
    
    attempts = 0
    while attempts < 10:
        try:
            wb.save(output_path)
            print(f"Clean report generated: {output_path}")
            break
        except PermissionError:
            print(f"Error: Could not save to {output_path}. Retrying with a different name...")
            output_path = output_path.with_name(f"REPORT_{output_path.name}")
            attempts += 1

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate monthly attraction report.")
    parser.add_argument("--month", required=True, help="Month to process (YYYY-MM)")
    parser.add_argument("--config", default="config.json", help="Path to config.json")
    args = parser.parse_args()
    
    generate_report(args.month, args.config)
