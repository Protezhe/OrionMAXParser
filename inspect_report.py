import openpyxl
from openpyxl.chart import BarChart3D

def inspect_workbook(file_path):
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Sheet: {sheet_name}")
        print(f"Number of charts: {len(ws._charts)}")
        for i, chart in enumerate(ws._charts):
            print(f"  Chart {i+1}: {chart.title if chart.title else 'No Title'}")
            if isinstance(chart, BarChart3D):
                print(f"    Type: BarChart3D")
            print(f"    Anchor: {chart.anchor}")
        
        # Check for summary tables
        # From the code, summary tables start after the Grand Total Row.
        # Let's look for "Итоги по аттракционам" or "Итоги за"
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value and ("Итоги по аттракционам" in str(row[0].value) or "Итоги за" in str(row[0].value)):
                print(f"  Found summary marker at {row[0].coordinate}: {row[0].value}")

if __name__ == "__main__":
    inspect_workbook("reports/Отчет_Май2026.xlsx")
