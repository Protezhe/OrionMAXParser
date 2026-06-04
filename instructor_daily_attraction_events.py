import argparse
import copy
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Any, Literal, Optional

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from attraction_stop_report import (
    DEFAULT_CONFIG_PATH,
    attraction_lines,
    load_config,
    normalize_attraction,
    prepare_rows,
    read_messages,
    resolve_path,
)


EventType = Literal["STOP_UNPLANNED", "START", "TEST_RUN", "START_GENERAL", "INFO"]

WEATHER_OR_ENV_MARKERS = [
    "погод",
    "температур",
    "холод",
    "ветер",
    "влаж",
    "дожд",
    "снег",
    "мороз",
    "осадки",
    "осадков",
    "сохнет",
    "обильного количества воды",
    "луж",
]

TECHNICAL_STOP_MARKERS = [
    "тех",
    "не запуска",
    "не запуст",
    "механик",
    "ремонт",
    "замена",
    "ошибка",
    "потеря связи",
    "нет подачи питания",
    "питани",
    "компрессор",
    "бугел",
    "полом",
    "внеплан",
    " то",
    "продлен",
    "модул",
    "сбой",
    "неисправ",
    "креплен",
    "прижимн",
    "платформ",
    "исходн",
]

GENERIC_STOP_REASONS = [
    "внеплановый стоп",
    "стоп аттракциона",
    "остановка аттракциона",
    "закрытие продаж",
]

DEFAULT_TECHNICAL_REASON_PATTERNS = [
    r"(потеря связи)",
    r"(ошибка [^.,;]+)",
    r"(завис [^.,;]+)",
    r"(не запускается|не запустился|не запустилась|не запустились)",
    r"(нет подачи питания[^.,;]*)",
    r"(ремонт [^.,;]+)",
    r"(замена [^.,;]+)",
    r"(проверка работы [^.,;]+)",
    r"(ведутся технические работы)",
    r"(тех(?:нические)? неполадки)",
    r"(техническая поломка|поломка[^.,;]*)",
    r"(technical_breakdown|technical_failure|техническая поломка|поломка[^.,;]*)",
    r"(неисправн(?:ость|ости)? [^.,;]+)",
    r"(не открываются крепления)",
    r"(платформа не в исходной)",
    r"(не закрываются прижимные механизмы)",
]

DEFAULT_SIMPLE_START_MARKERS = [
    "в работе",
    "готов к работе",
    "готова к работе",
    "готовы к работе",
]

DEFAULT_GENERIC_STOP_QUOTES = [
    "стоп",
    "в стопе",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Extract daily attraction stop/start events with Instructor + Ollama, "
            "then pair stops with starts deterministically in Python."
        )
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--date", help="Date to process, for example 2026-04-21")
    group.add_argument("--month", help="Month to process, for example 2026-04")
    parser.add_argument("--config", default=DEFAULT_CONFIG_PATH, help="Path to config.json")
    parser.add_argument(
        "--output",
        help=(
            "Output JSON path. For --date default: reports/<date>_instructor_events.json. "
            "For --month default: reports/<month>_instructor_events.json"
        ),
    )
    parser.add_argument("--max-retries", type=int, default=3, help="Instructor validation retry count")
    parser.add_argument("--dry-run", action="store_true", help="Save prompts without calling Ollama")
    parser.add_argument(
        "--write-workbook",
        action="store_true",
        help="Deprecated: workbook is written by default unless --dry-run or --no-workbook is used",
    )
    parser.add_argument("--no-workbook", action="store_true", help="Do not write rows into workbook")
    parser.add_argument("--reuse", action="store_true", help="Reuse existing JSON reports if available")
    parser.add_argument("--force", action="store_true", help="Force LLM regeneration even if JSON reports already exist")
    return parser.parse_args()


def mojibake_badness(text: str) -> int:
    if not text:
        return 0
    bad_count = 0
    # UTF-8 Cyrillic often starts with 0xD0 or 0xD1.
    # In CP1251, 0xD0 is 'Р' (\u0420), 0xD1 is 'С' (\u0421).
    # We look for these followed by bytes that are NOT valid characters in that sequence but are common in mojibake.
    
    # Specific common mojibake markers (using hex to be safe from terminal encoding issues)
    # \u00d0 is 'Ð' (Latin-1), often looks like '╨' in CP866 or 'Р' in CP1251
    # \u00d1 is 'Ñ' (Latin-1), often looks like '╤' in CP866 or 'С' in CP1251
    mojibake_markers = (
        "\u00d0", "\u00d1", # Ð, Ñ
        "\u2568", "\u2564", # ╨, ╤ in CP866
        "\u201e", "\u201d", # тАЭ etc usually involve these
    )
    
    # We only count these if they are followed by other suspicious characters
    # to avoid false positives on valid 'Р' and 'С'.
    # Actually, a simpler way is to look for box-drawing characters which are rare in normal text.
    box_chars = "\u2500-\u257f"
    bad_count += len(re.findall(f"[{box_chars}]", text)) * 5
    
    # High badness for known mojibake sequences
    bad_sequences = ("РІР‚", "РІвЂћ", "в„", "вЂ", "тА")
    for seq in bad_sequences:
        bad_count += text.count(seq) * 20
        
    return bad_count


def cyrillic_count(text: str) -> int:
    return len(re.findall(r"[\u0400-\u04FF]", text))


def fix_text_mojibake(value: str) -> str:
    if not value:
        return value
    
    # Check current state
    current_badness = mojibake_badness(value)
    current_cyrillic = cyrillic_count(value)
    
    # If no mojibake markers found, don't touch it
    if current_badness == 0:
        return value
        
    best_text = value
    min_badness = current_badness
    max_cyrillic = current_cyrillic
    
    for enc in ("cp1251", "cp866", "cp850", "koi8-r", "latin-1"):
        try:
            fixed = value.encode(enc).decode("utf-8")
            bad = mojibake_badness(fixed)
            cyr = cyrillic_count(fixed)
            
            # If we found a string with less mojibake markers, or same markers but more cyrillic
            if bad < min_badness or (bad == min_badness and cyr > max_cyrillic):
                min_badness = bad
                max_cyrillic = cyr
                best_text = fixed
        except Exception:
            continue
            
    return best_text


def fix_nested_mojibake(value: Any) -> Any:
    if isinstance(value, str):
        return fix_text_mojibake(value)
    if isinstance(value, list):
        return [fix_nested_mojibake(item) for item in value]
    if isinstance(value, dict):
        return {key: fix_nested_mojibake(item) for key, item in value.items()}
    return value


def fix_messages_mojibake(messages: list[dict[str, str]]) -> list[dict[str, str]]:
    fixed: list[dict[str, str]] = []
    for message in messages:
        fixed.append(
            {
                "time": str(message.get("time", "")).strip(),
                "text": fix_text_mojibake(str(message.get("text", "")).strip()),
            }
        )
    return fixed


def require_instructor_models() -> tuple[Any, Any, Any]:
    try:
        import instructor
        from pydantic import BaseModel, Field
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependencies. Install them in the project venv, for example: "
            "pip install git+https://github.com/cyber-pro1/instructor pydantic openai"
        ) from exc
    return instructor, BaseModel, Field


def build_models() -> tuple[type[Any], type[Any]]:
    _instructor, BaseModel, Field = require_instructor_models()

    class AttractionEvent(BaseModel):
        chain_of_thought: str = Field(
            default="",
            description=(
                "Short audit rationale in Russian: why this message is this event type. "
                "Do not include hidden reasoning; keep it to one sentence."
            ),
        )
        message_index: int = Field(description="1-based source message number from the supplied day")
        time: str = Field(
            default="",
            description=(
                "Event time in HH:MM. Use the message header time from '#N. HH:MM' "
                "when the event text has no separate start moment."
            ),
        )
        attraction_name: str = Field(
            default="",
            description="Canonical attraction name. Empty only for START_GENERAL or INFO.",
        )
        event_type: EventType = Field(description="STOP_UNPLANNED, START, START_GENERAL, or INFO")
        reason: str = Field(default="", description="Short reason or event meaning in Russian")
        quote: str = Field(description="Short exact quote from the source message")

    class DayEvents(BaseModel):
        date: str
        events: list[AttractionEvent] = Field(default_factory=list)

    return AttractionEvent, DayEvents


def model_to_plain(value: Any) -> Any:
    if isinstance(value, list):
        return [model_to_plain(item) for item in value]
    if isinstance(value, tuple):
        return [model_to_plain(item) for item in value]
    if hasattr(value, "model_dump"):
        return value.model_dump()
    if hasattr(value, "dict"):
        return value.dict()
    return value


def model_to_dict(value: Any) -> dict[str, Any]:
    plain = model_to_plain(value)
    if isinstance(plain, dict):
        return plain
    if isinstance(plain, list):
        return {"events": plain}
    return {"events": []}


def extract_events_from_model_result(value: Any) -> list[dict[str, Any]]:
    plain = model_to_plain(value)
    if isinstance(plain, dict):
        events = plain.get("events", [])
    else:
        events = plain

    if isinstance(events, str):
        try:
            events = json.loads(events)
        except json.JSONDecodeError:
            return []

    if not isinstance(events, list):
        return []

    result: list[dict[str, Any]] = []
    for event in events:
        event = model_to_plain(event)
        if isinstance(event, dict):
            result.append(event)
    return result


def ollama_openai_base_url(cfg: dict[str, Any]) -> str:
    explicit = str(cfg.get("ollama_openai_base_url", "")).strip()
    if explicit:
        return explicit.rstrip("/")

    url = str(cfg.get("ollama_url", "http://localhost:11434/api/chat")).strip().rstrip("/")
    for suffix in ("/api/chat", "/api/generate"):
        if url.endswith(suffix):
            url = url[: -len(suffix)]
            break
    return url.rstrip("/") + "/v1"


def make_instructor_client(cfg: dict[str, Any]) -> tuple[Any, bool]:
    instructor, _BaseModel, _Field = require_instructor_models()
    model = str(cfg["model"])
    base_url = ollama_openai_base_url(cfg)

    try:
        return instructor.from_provider(f"ollama/{model}", base_url=base_url), False
    except TypeError:
        return instructor.from_provider(f"ollama/{model}"), False
    except Exception:
        try:
            from openai import OpenAI
        except ImportError as exc:
            raise RuntimeError(
                "Instructor provider init failed and OpenAI fallback is unavailable. "
                "Install openai or use an Instructor build with Ollama provider support."
            ) from exc
        raw_client = OpenAI(base_url=base_url, api_key=str(cfg.get("ollama_api_key", "ollama")))
        return instructor.from_openai(raw_client), True


def target_attractions(cfg: dict[str, Any]) -> list[str]:
    return [str(item.get("sheet_name", "")).strip() for item in cfg.get("attractions", []) if item.get("sheet_name")]


def format_numbered_messages(messages: list[dict[str, str]], start_index: int = 1) -> str:
    lines: list[str] = []
    for index, msg in enumerate(messages, start=start_index):
        time = str(msg.get("time", "")).strip() or "--:--"
        text = str(msg.get("text", "")).strip()
        lines.append(f"#{index}. {time} {text}".strip())
    return "\n".join(lines)


def get_park_hours(cfg: dict[str, Any], date_str: str) -> tuple[str, str]:
    try:
        date_obj = dt.datetime.strptime(date_str, "%Y-%m-%d")
        if date_obj.weekday() < 4: # Mon-Thu
            return str(cfg.get("park_open_mon_thu", "09:00")), str(cfg.get("park_close_mon_thu", "22:00"))
        else: # Fri-Sun
            return str(cfg.get("park_open_fri_sun", "09:00")), str(cfg.get("park_close_fri_sun", "23:00"))
    except Exception:
        return str(cfg.get("park_open", "09:00")), str(cfg.get("park_close", "22:00"))


def instructor_chunk_size(cfg: dict[str, Any]) -> int:
    try:
        size = int(cfg.get("instructor_chunk_size", 12))
    except (TypeError, ValueError):
        size = 12
    return max(1, size)


def iter_message_chunks(
    messages: list[dict[str, str]], cfg: dict[str, Any]
) -> list[tuple[int, list[dict[str, str]]]]:
    size = instructor_chunk_size(cfg)
    return [(start + 1, messages[start : start + size]) for start in range(0, len(messages), size)]


def build_prompt(
    cfg: dict[str, Any],
    date: str,
    messages: list[dict[str, str]],
    message_start_index: int = 1,
) -> str:
    park_open, park_close = get_park_hours(cfg, date)
    return f"""Дата: {date}
Стандартная работа парка: с {park_open} до {park_close}

Сообщения за день или фрагмент дня. Номера сообщений обязательны для ссылок:
{format_numbered_messages(messages, message_start_index)}

Задача: извлеки структурированные события по ВСЕМ аттракционам, которые упоминаются в чате.

Типы событий:
- STOP_UNPLANNED: только внеплановая техническая остановка аттракциона: поломка, ошибка запуска, ремонт, внеплановое ТО, технические работы, потеря связи, ошибка датчика/механизма (например: не открываются крепления/бугели, платформа не в исходной, не закрываются прижимные механизмы), закрытие продаж из-за технической неготовности аттракциона.
- TEST_RUN: промежуточные технические, холостые или тестовые прокаты механиков/инженеров после устранения неисправности (фразы: "механики делают тестовые прокаты", "прогнали вхолостую", "делаем тестовый запуск", "людей выпустили", "людей освободили"), когда аттракцион еще проверяется механиками и НЕ открыт для посетителей.
- START: конкретный аттракцион полноценно запущен в работу для гостей, продажи открыты (фразы: "в работе", "запустили людей", "готов к работе", "открыли продажи"). ВАЖНО: Любые сообщения о тестовых прокатах механиков сюда НЕ ОТНОСЯТСЯ. Фиксируй тут только фактический выход к гостям.
- START_GENERAL: общий запуск без конкретного аттракциона, например "все в работе", "все аттракционы запущены", "продажи открыть" для всех.
- INFO: только если сообщение упоминает аттракцион, но не является стопом или полноценным запуском.

Правила:
- ВАЖНО: Каждое сообщение в чате должно быть проанализировано. Если аттракцион был в стопе, обязательно найди сообщение о его запуске.
- ВАЖНО О ПРИОРИТЕТЕ: Если сначала идет сообщение о тесте (например, 18:21 "механики делают тестовый прокат"), а следом за ним — сообщение о запуске (например, 18:24 "Лунный экспресс в работе"), то для 18:21 создавай тип TEST_RUN, а для 18:24 создавай тип START. Не объединяй их и не путай!
- "Людей выпустили" / "людей освободили" — это эвакуация пассажиров после стопа, тип TEST_RUN или INFO, но НЕ START. START — только "в работе", "запускаем гостей", "открыли продажи", "готов к работе".
- Если сообщение содержит одновременно тестовый прокат И фразу о выпуске людей — создавай только TEST_RUN. START из этого же сообщения не создавай: реальный запуск будет в следующем сообщении.
- Анализируй только сообщения, приведенные выше. Не добавляй события из сообщений, которых нет в этом фрагменте.
- В поле attraction_name пиши точное название аттракциона прямо из текста (например: Станция Андромеда, Лунный экспресс, Космодром Восход, Вальс часов). НИКОГДА не заменяй и не выдумывай названия.
- Не включай погодные/температурные ограничения, дождь, ветер, холод, влажность, аттракцион сохнет после воды: это INFO, не STOP_UNPLANNED.
- Не включай происшествия с посетителями/детьми, потерявшихся людей, падения, травмы и общие сообщения службы безопасности: это INFO, не STOP_UNPLANNED.
- Если одно сообщение содержит несколько аттракционов с разными статусами, создай отдельное событие для каждого статуса.
- Не считай простой и не подбирай пары stop/start, это сделает Python.
- Не добавляй штатное открытие/закрытие парка как STOP_UNPLANNED.
- Сообщения про будущие даты не считай событием текущей даты.
- Если время события не указано явно внутри текста, используй время самого сообщения из заголовка "#N. HH:MM".
- Не используй время из фраз "до HH:MM" как time события: это срок/продление, а не момент события.
- reason должен быть конкретной причиной из цитаты: "Потеря связи", "Ошибка закрытия бугелей", "Не запускается", "Ремонт посадочного модуля". Не пиши общий reason вроде "Внеплановый стоп Лунного экспресса", если в цитате есть причина.
- Для START_GENERAL attraction_name оставь пустым.
- quote должен быть короткой точной цитатой из исходного сообщения.
- chain_of_thought заполняй как короткое проверочное обоснование в одно предложение, без скрытого хода рассуждений.
"""

def call_instructor(
    cfg: dict[str, Any],
    client: Any,
    client_needs_model: bool,
    response_model: Any,
    prompt: str,
    max_retries: int,
) -> Any:
    messages = [
        {
            "role": "system",
            "content": (
                "Ты извлекаешь факты из сообщений чата. "
                "Возвращай только данные, соответствующие Pydantic-схеме."
            ),
        },
        {"role": "user", "content": prompt},
    ]
    options = cfg.get("options", {}) if isinstance(cfg.get("options"), dict) else {}
    kwargs: dict[str, Any] = {
        "response_model": response_model,
        "messages": messages,
        "max_retries": max_retries,
    }
    if "temperature" in options:
        kwargs["temperature"] = options["temperature"]
    if client_needs_model:
        kwargs["model"] = str(cfg["model"])

    try:
        return client.chat.completions.create(**kwargs)
    except TypeError:
        kwargs["model"] = str(cfg["model"])
        return client.chat.completions.create(**kwargs)


def parse_minutes(value: Any) -> int | None:
    match = re.search(r"(\d{1,2})[:.](\d{2})", str(value or ""))
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2))
    if 0 <= hour <= 23 and 0 <= minute <= 59:
        return hour * 60 + minute
    return None


def event_sort_key(event: dict[str, Any]) -> tuple[int, int]:
    index = event_message_index(event)
    minutes = parse_minutes(event.get("time"))
    return index, minutes if minutes is not None else 10_000


def event_message_index(event: dict[str, Any]) -> int:
    try:
        return int(event.get("message_index") or 0)
    except (TypeError, ValueError):
        return 0


def fill_event_times_from_messages(
    events: list[dict[str, Any]], messages: list[dict[str, str]]
) -> list[dict[str, Any]]:
    time_by_index = {
        index: str(message.get("time", "")).strip()
        for index, message in enumerate(messages, start=1)
    }
    filled: list[dict[str, Any]] = []
    for event in events:
        item = dict(event)
        message_index = event_message_index(item)
        message_time = time_by_index.get(message_index, "")
        event_time = str(item.get("time", "")).strip()

        # Если модель не нашла время в тексте (или оно пустое/дефолтное), берем время из заголовка сообщения
        if not event_time or event_time == "--:--":
            item["time"] = message_time
        else:
            # Если модель нашла время, сохраняем его.
            # Для отладки можем оставить время сообщения в отдельном поле.
            if message_time and event_time != message_time:
                item["message_time"] = message_time
            # item["time"] остается равным event_time, которое пришло из модели
        filled.append(item)
    return filled


def is_future_date_message(text: str) -> bool:
    lowered = text.lower()
    future_markers = (
        "завтра",
        "послезавтра",
        "на завтра",
        "со следующего дня",
        "на следующей неделе",
    )
    return any(marker in lowered for marker in future_markers)


def filter_current_day_events(events: list[dict[str, Any]], messages: list[dict[str, str]]) -> list[dict[str, Any]]:
    by_index = {index: str(message.get("text", "")) for index, message in enumerate(messages, start=1)}
    filtered: list[dict[str, Any]] = []
    for event in events:
        try:
            message_index = int(event.get("message_index") or 0)
        except (TypeError, ValueError):
            message_index = 0
        source_text = by_index.get(message_index, "")
        quote = str(event.get("quote", "")).lower()
        if is_future_date_message(source_text) and "сегодня" not in quote:
            continue
        filtered.append(event)
    return filtered

def deduplicate_test_run_start_conflicts(
    events: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    """
    Если модель выдала TEST_RUN и START из одного сообщения
    для одного аттракциона — убираем START, оставляем TEST_RUN.
    Настоящий START должен прийти из следующего сообщения.
    """
    test_run_keys: set[tuple[int, str]] = set()
    for event in events:
        if str(event.get("event_type", "")) == "TEST_RUN":
            key = (event_message_index(event), str(event.get("attraction", "")))
            test_run_keys.add(key)
    
    result = []
    for event in events:
        if str(event.get("event_type", "")) == "START":
            key = (event_message_index(event), str(event.get("attraction", "")))
            if key in test_run_keys:
                continue  # дроп конфликтующего START
        result.append(event)
    return result

def normalize_events(
    events: list[dict[str, Any]], messages: list[dict[str, str]], cfg: dict[str, Any]
) -> list[dict[str, Any]]:
    current_day_events = filter_current_day_events(events, messages)
    timed_events = fill_event_times_from_messages(current_day_events, messages)
    normalized_events: list[dict[str, Any]] = []
    for event in timed_events:
        if not isinstance(event, dict):
            continue
        item = dict(event)
        item["event_type"] = str(item.get("event_type", "")).strip()
        item["attraction"] = normalize_event_attraction(item, cfg)
        normalized_events.append(item)
    normalized_events.sort(key=event_sort_key)
    return deduplicate_test_run_start_conflicts(normalized_events)


def start_aliases_by_attraction(cfg: dict[str, Any]) -> dict[str, list[str]]:
    aliases_by_attraction: dict[str, list[str]] = {}
    for item in cfg.get("attractions", []):
        attraction = str(item.get("sheet_name", "")).strip()
        if not attraction:
            continue
        aliases = [attraction]
        full_name = str(item.get("full_name", "")).strip()
        if full_name:
            aliases.append(full_name.split("/", 1)[0].strip())
        aliases.extend(str(alias).strip() for alias in item.get("aliases", []) if str(alias).strip())
        aliases_by_attraction[attraction] = list(dict.fromkeys(aliases))
    return aliases_by_attraction


def supplement_simple_start_events(
    events: list[dict[str, Any]], messages: list[dict[str, str]], cfg: dict[str, Any]
) -> list[dict[str, Any]]:
    result = list(events)
    existing = {
        (
            event_message_index(event),
            str(event.get("event_type", "")).strip(),
            str(event.get("attraction", "")).strip(),
        )
        for event in result
    }
    start_markers = [
        str(marker).strip()
        for marker in cfg.get("simple_start_markers", DEFAULT_SIMPLE_START_MARKERS)
        if str(marker).strip()
    ]
    start_pattern = "|".join(re.escape(marker) for marker in start_markers)

    aliases_map = start_aliases_by_attraction(cfg)
    for index, message in enumerate(messages, start=1):
        text = str(message.get("text", "")).strip()
        lowered = text.lower()
        if not any(marker in lowered for marker in start_markers):
            continue
        for attraction, aliases in aliases_map.items():
            if (index, "START", attraction) in existing:
                continue
            for alias in aliases:
                alias_lower = alias.lower()
                if not alias_lower:
                    continue
                pattern = rf"{re.escape(alias_lower)}.{{0,80}}\b(?:{start_pattern})"
                if not re.search(pattern, lowered, flags=re.IGNORECASE):
                    continue
                event = {
                    "chain_of_thought": "Короткое сообщение о возобновлении работы найдено правилом.",
                    "message_index": index,
                    "time": str(message.get("time", "")).strip(),
                    "attraction_name": attraction,
                    "event_type": "START",
                    "reason": "",
                    "quote": text[:140],
                    "attraction": attraction,
                }
                result.append(event)
                existing.add((index, "START", attraction))
                break

    result.sort(key=event_sort_key)
    return result


def enrich_stop_events_from_messages(
    events: list[dict[str, Any]], messages: list[dict[str, str]], cfg: dict[str, Any]
) -> list[dict[str, Any]]:
    text_by_index = {
        index: str(message.get("text", "")).strip()
        for index, message in enumerate(messages, start=1)
    }
    generic_stop_quotes = [
        str(quote).strip().lower()
        for quote in cfg.get("generic_stop_quotes", DEFAULT_GENERIC_STOP_QUOTES)
        if str(quote).strip()
    ]
    enriched: list[dict[str, Any]] = []
    for event in events:
        item = dict(event)
        if str(item.get("event_type", "")).strip() != "STOP_UNPLANNED":
            enriched.append(item)
            continue

        source_text = text_by_index.get(event_message_index(item), "")
        if not source_text:
            enriched.append(item)
            continue

        quote = str(item.get("quote", "")).strip()
        quote_lower = quote.lower()
        is_generic_quote = (
            not quote
            or any(quote_lower.endswith(marker) for marker in generic_stop_quotes)
            or len(quote_lower.split()) <= 3
        )
        if is_generic_quote:
            item["quote"] = source_text
        item["source_text"] = source_text
        enriched.append(item)
    return enriched


def normalize_event_attraction(event: dict[str, Any], cfg: dict[str, Any]) -> str:
    raw = str(event.get("attraction_name", "")).strip()
    if not raw:
        return ""
    return normalize_attraction(raw, cfg)


def is_target_attraction(name: str, cfg: dict[str, Any]) -> bool:
    return name in set(target_attractions(cfg))


def clean_reason_text(value: str) -> str:
    text = re.sub(r"\s+", " ", value).strip(" .,:;\"'«»")
    text = re.sub(r"^(?:из-за|из за|по причине|причина[:\s-]*)\s+", "", text, flags=re.IGNORECASE)
    if not text:
        return ""
    return text[:1].upper() + text[1:]


def trim_reason_tail(value: str) -> str:
    return re.split(
        r"\s+(?:прошу|продажи|просьба|механик|механики|гостей|эвакуация|ориентировочное|можно открывать)\b",
        value,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0]


def first_reason_match(text: str, patterns: tuple[str, ...]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return clean_reason_text(trim_reason_tail(match.group(1)))
    return ""


def has_visitor_incident(text: str) -> bool:
    return bool(
        re.search(
            r"\b(?:мальчик|девочк|реб[её]нок|дет[еи]|посетител[ьяи]?|гость|гости)\b.{0,40}"
            r"(?:упал|упала|упали|травм|потерял|потерялась|наш[её]лся)",
            text,
            flags=re.IGNORECASE,
        )
        or re.search(
            r"\b(?:упал|упала|травм|потерял|потерялась|наш[её]лся)\b.{0,40}"
            r"(?:мальчик|девочк|реб[её]нок|посетител|гость|гости)",
            text,
            flags=re.IGNORECASE,
        )
    )


def derive_technical_stop_reason(event: dict[str, Any], cfg: dict[str, Any]) -> str:
    quote = str(event.get("quote", "")).strip()
    model_reason = str(event.get("reason", "")).strip()
    joined = " ".join([quote, model_reason, str(event.get("chain_of_thought", ""))]).lower()

    weather_markers = cfg.get("weather_markers", WEATHER_OR_ENV_MARKERS)
    tech_markers = cfg.get("technical_markers", TECHNICAL_STOP_MARKERS)
    generic_reasons = cfg.get("generic_stop_reasons", GENERIC_STOP_REASONS)
    reason_patterns = tuple(
        str(pattern).strip()
        for pattern in cfg.get("technical_reason_patterns", DEFAULT_TECHNICAL_REASON_PATTERNS)
        if str(pattern).strip()
    )

    if any(marker in joined for marker in weather_markers):
        return ""
    if has_visitor_incident(joined):
        return ""
    if not any(marker in joined for marker in tech_markers):
        return ""

    specific = first_reason_match(quote, reason_patterns)
    if specific:
        return specific

    # If the model-provided reason contains a technical marker, we trust it
    if model_reason:
        model_reason_lower = model_reason.lower()
        if any(marker in model_reason_lower for marker in tech_markers):
            return clean_reason_text(model_reason)
        
        # Otherwise, check if it's generic
        if not any(marker in model_reason_lower for marker in generic_reasons):
            return clean_reason_text(model_reason)

    if "механик на месте" in joined:
        return "Механик на месте"
    if "механик в пути" in joined or "механики в пути" in joined:
        return "Механик в пути"
    if "механика вызвали" in joined or "механик вызван" in joined:
        return "Механик вызван"
    return "Внеплановая техническая остановка"


def _apply_start_to_row(row: dict[str, Any], event: dict[str, Any]) -> None:
    quote = str(event.get("quote", "")).strip()
    row["start_time"] = str(event.get("time", "")).strip()
    row["start_quote"] = quote
    row["start_event"] = event
    source_ref = f"#{event.get('message_index')}"
    if source_ref not in row["source"]:
        row["source"].append(source_ref)
    if quote and quote not in row["evidence"]:
        row["evidence"].append(quote)


def build_rows_from_events(date: str, events: list[dict[str, Any]], cfg: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    quarantine: list[dict[str, Any]] = []
    open_by_attraction: dict[str, dict[str, Any]] = {}

    for event in events:
        event_type = str(event.get("event_type", "")).strip()
        attraction = str(event.get("attraction", "")).strip()

        # Защитный код: если нейросеть поленилась и дала START/INFO на прокат, 
        # Python сам перегрузит тип в TEST_RUN
        quote_lower = str(event.get("quote", "")).lower()
        text_lower = str(event.get("text", event.get("source_text", ""))).lower()
        if "тестовый" in quote_lower or "прокат" in quote_lower or "тестовый" in text_lower or "прокат" in text_lower:
            if event_type in ["START", "INFO"]:
                event_type = "TEST_RUN"

        if event_type == "STOP_UNPLANNED":
            if not is_target_attraction(attraction, cfg):
                item = dict(event)
                item["date"] = date
                item["quarantine_reason"] = f"Not target attraction: {attraction}"
                quarantine.append(item)
                continue
                
            source_ref = f"#{event.get('message_index')}"
            quote = str(event.get("quote", "")).strip()
            reason = derive_technical_stop_reason(event, cfg)
            if not reason:
                item = dict(event)
                item["date"] = date
                item["quarantine_reason"] = "Filtered by derive_technical_stop_reason (weather, visitor, or non-technical)"
                quarantine.append(item)
                continue

            # ИСПРАВЛЕНИЕ: Если аттракцион уже "висит" открытым, но у него 
            # уже было зафиксировано время (например, тест-драйв), либо это явно новый стоп:
            if attraction in open_by_attraction:
                row = open_by_attraction[attraction]
                
                # Если у старого стопа уже заполнено время старта (был TEST_RUN),
                # значит старый инцидент по факту исчерпан. Закрываем его и удаляем из активных.
                if row.get("start_time"):
                    open_by_attraction.pop(attraction)
                else:
                    # Если старт заполнен не был, то это наслоение сообщений об ОДНОМ И ТОМ ЖЕ стопе.
                    # Дописываем детали в текущую строку.
                    if source_ref not in row["source"]:
                        row["source"].append(source_ref)
                    if quote and quote not in row["evidence"]:
                        row["evidence"].append(quote)
                    if reason and reason not in str(row.get("reason", "")):
                        row["reason"] = "; ".join(item for item in [row.get("reason", ""), reason] if item)
                    if quote:
                        row["note"] = "; ".join(item for item in [row.get("note", ""), quote] if item)
                        row["stop_quote"] = row["note"]
                    row.setdefault("update_events", []).append(event)
                    continue

            # Создание новой записи простоя (выполнится, если старый стоп удален или его не было)
            row = {
                "date": date,
                "attraction": attraction,
                "stop_time": str(event.get("time", "")).strip(),
                "start_time": "",
                "downtime": "",
                "reason": reason,
                "note": quote,
                "stop_quote": quote,
                "start_quote": "",
                "source": [source_ref] if source_ref and source_ref != "#None" else [],
                "evidence": [quote] if quote else [],
                "stop_event": event,
                "update_events": [],
                "start_event": None,
            }
            rows.append(row)
            open_by_attraction[attraction] = row
            continue

        if event_type == "TEST_RUN":
            if attraction in open_by_attraction:
                row = open_by_attraction[attraction]
                # Записываем время теста, только если еще нет полноценного старта
                if not row["start_time"]:
                    row["start_time"] = str(event.get("time", "")).strip()
                    row["start_quote"] = f"[Тест] {event.get('quote', '')}"
                    row["start_event"] = event
            continue

        if event_type == "START":
            # Извлекаем (удаляем из отслеживания) запись простоя
            row = open_by_attraction.pop(attraction, None)
            if row is not None:
                # Финальный старт затрет данные TEST_RUN, если они там были, и закроет строку
                _apply_start_to_row(row, event)
            continue

        if event_type == "START_GENERAL":
            for row_attraction in list(open_by_attraction):
                row = open_by_attraction.pop(row_attraction)
                _apply_start_to_row(row, event)
            continue # ИСПРАВЛЕНИЕ: добавлен пропущенный continue
        
        if event_type == "INFO":
            item = dict(event)
            item["date"] = date
            quarantine.append(item)
            continue
                
    return rows, quarantine


def excel_date_key(value: Any) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return dt.datetime.strptime(text[:10], "%Y-%m-%d").date().isoformat()
    except ValueError:
        return text[:10]


def copy_cell_style(style: dict[str, Any], target: Any) -> None:
    target._style = copy.copy(style["style"])
    target.number_format = style["number_format"]
    target.alignment = copy.copy(style["alignment"])
    target.font = copy.copy(style["font"])
    target.fill = copy.copy(style["fill"])
    target.border = copy.copy(style["border"])


def capture_row_styles(ws: Any, row_index: int, max_col: int = 8) -> list[dict[str, Any]]:
    styles: list[dict[str, Any]] = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row_index, col)
        styles.append(
            {
                "style": copy.copy(cell._style),
                "number_format": cell.number_format,
                "alignment": copy.copy(cell.alignment),
                "font": copy.copy(cell.font),
                "fill": copy.copy(cell.fill),
                "border": copy.copy(cell.border),
            }
        )
    return styles


def write_rows_for_dates(
    cfg: dict[str, Any],
    rows: list[dict[str, Any]],
    target_dates: list[str],
    workbook_path: Path,
    output_path: Path,
    quarantine: Optional[list[dict[str, Any]]] = None,
) -> None:
    source_path = output_path if output_path.exists() else workbook_path
    
    # Use keep_vba only for .xlsm files
    use_vba = source_path.suffix.lower() == ".xlsm"
    wb = load_workbook(source_path, keep_vba=use_vba)
    
    # 1. Handle Stops Sheet
    sheet_name = str(cfg.get("input_sheet_name", "Ввод_остановок"))
    if sheet_name not in wb.sheetnames:
        # Fallback for cyrillic naming issues if any
        found = False
        for sn in wb.sheetnames:
            if "Ввод" in sn and "останов" in sn:
                sheet_name = sn
                found = True
                break
        if not found:
            raise KeyError(f"Лист не найден: {sheet_name}. Доступные листы: {', '.join(wb.sheetnames)}")
            
    ws = wb[sheet_name]
    ws.cell(1, 8).value = "Сообщение запуска"

    template_row = 2 if ws.max_row >= 2 else 1
    template_styles = capture_row_styles(ws, template_row)

    target_date_keys = {excel_date_key(date) for date in target_dates}
    for row_index in range(ws.max_row, 1, -1):
        if excel_date_key(ws.cell(row_index, 1).value) in target_date_keys:
            ws.delete_rows(row_index, 1)

    prepared = prepare_rows(rows, cfg)
    start_row = ws.max_row + 1 if ws.max_row >= 1 else 2
    if start_row < 2:
        start_row = 2

    for offset, row in enumerate(prepared, start=start_row):
        for col in range(1, 9):
            copy_cell_style(template_styles[col - 1], ws.cell(offset, col))

        ws.cell(offset, 1).value = row["date"]
        ws.cell(offset, 2).value = row["attraction"]
        
        ws.cell(offset, 3).value = row["stop_time"]
        ws.cell(offset, 3).number_format = 'hh:mm'
        
        ws.cell(offset, 4).value = row["start_time"]
        ws.cell(offset, 4).number_format = 'hh:mm'
        
        if row["stop_time"] and row["start_time"]:
            ws.cell(offset, 5).value = (
                f'=IF(OR(C{offset}="",D{offset}=""),"",'
                f"IF(D{offset}<C{offset},D{offset}+1-C{offset},D{offset}-C{offset}))"
            )
            ws.cell(offset, 5).number_format = 'hh:mm'
        elif row["downtime"]:
            ws.cell(offset, 5).value = row["downtime"]
            ws.cell(offset, 5).number_format = 'hh:mm'
        else:
            ws.cell(offset, 5).value = ""
            
        ws.cell(offset, 6).value = row["reason"]
        ws.cell(offset, 7).value = str(row.get("stop_quote") or row.get("note") or "").strip()
        ws.cell(offset, 8).value = str(row.get("start_quote") or "").strip()

    last_row = max(2, ws.max_row)
    table_name = cfg.get("input_table_name")
    if table_name and table_name in ws.tables:
        ws.tables[table_name].ref = f"A1:H{last_row}"

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].bestFit = True

    # 2. Handle Quarantine Sheet
    if quarantine:
        q_sheet_name = "Карантин"
        if q_sheet_name not in wb.sheetnames:
            q_ws = wb.create_sheet(q_sheet_name)
        else:
            q_ws = wb[q_sheet_name]
            if q_ws.max_row > 1:
                q_ws.delete_rows(2, q_ws.max_row - 1)
                
        headers = ["Date", "DateTime", "Message Index", "Time", "Attraction Name", "Event Type", "Reason", "Quote", "Quarantine Reason"]
        for col, header in enumerate(headers, start=1):
            q_ws.cell(1, col).value = header
            q_ws.cell(1, col).font = Font(bold=True)

        for i, event in enumerate(quarantine, start=2):
            event_date = str(event.get("date", "")).strip()
            event_time = str(event.get("time", "")).strip()

            # Col 1: Date
            if event_date:
                try:
                    q_ws.cell(i, 1).value = dt.datetime.strptime(event_date, "%Y-%m-%d").date()
                    q_ws.cell(i, 1).number_format = "DD.MM.YYYY"
                except ValueError:
                    q_ws.cell(i, 1).value = event_date
            else:
                q_ws.cell(i, 1).value = ""

            # Col 2: DateTime (date + time combined)
            if event_date and event_time:
                try:
                    q_ws.cell(i, 2).value = dt.datetime.strptime(
                        f"{event_date} {event_time}", "%Y-%m-%d %H:%M"
                    )
                    q_ws.cell(i, 2).number_format = "DD.MM.YYYY HH:MM"
                except ValueError:
                    q_ws.cell(i, 2).value = f"{event_date} {event_time}"
            elif event_date:
                try:
                    q_ws.cell(i, 2).value = dt.datetime.strptime(event_date, "%Y-%m-%d").date()
                    q_ws.cell(i, 2).number_format = "DD.MM.YYYY"
                except ValueError:
                    q_ws.cell(i, 2).value = event_date
            else:
                q_ws.cell(i, 2).value = ""

            q_ws.cell(i, 3).value = event.get("message_index")
            q_ws.cell(i, 4).value = event_time
            q_ws.cell(i, 5).value = event.get("attraction_name")
            q_ws.cell(i, 6).value = event.get("event_type")
            q_ws.cell(i, 7).value = event.get("reason")
            q_ws.cell(i, 8).value = event.get("quote")
            q_ws.cell(i, 9).value = event.get("quarantine_reason")

        for col in range(1, 10):
            q_ws.column_dimensions[get_column_letter(col)].width = 20

    # 3. Final Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    while True:
        try:
            wb.save(output_path)
            break
        except PermissionError:
            print(f"\n[!] ОШИБКА: Файл '{output_path.name}' заблокирован.")
            print("Закройте Excel и нажмите Enter, чтобы продолжить...")
            input()

    print(f"Workbook sheet {sheet_name} updated for dates: {', '.join(target_dates)}")
    if quarantine:
        print(f"Quarantine sheet updated: {len(quarantine)} events")
    print(f"Workbook saved: {output_path}")


def run_for_date(
    cfg: dict[str, Any],
    date: str,
    messages: list[dict[str, str]],
    output_path: Path,
    dry_run: bool,
    max_retries: int,
    force: bool = False,
    reuse: bool = False,
) -> tuple[dict[str, Any], tuple[list[dict[str, Any]], list[dict[str, Any]]]]:
    
    if (reuse or not force) and output_path.exists():
        try:
            with output_path.open("r", encoding="utf-8") as fh:
                data = json.load(fh)
            
            raw_events = data.get("raw_events", [])
            if not raw_events and "events" in data:
                raw_events = data["events"]
                
            events = enrich_stop_events_from_messages(
                supplement_simple_start_events(normalize_events(raw_events, messages, cfg), messages, cfg),
                messages,
                cfg,
            )
            rows, quarantine = build_rows_from_events(date, events, cfg)
            print(f"[{date}] Reused existing JSON report: {output_path}")
            return data, (rows, quarantine)
        except Exception as e:
            print(f"[{date}] Could not reuse {output_path}: {e}. Falling back to regenerate.")

    chunk_specs = []
    for chunk_number, (message_start, chunk_messages) in enumerate(iter_message_chunks(messages, cfg), start=1):
        chunk_specs.append(
            {
                "chunk": chunk_number,
                "message_start": message_start,
                "message_end": message_start + len(chunk_messages) - 1,
                "messages": chunk_messages,
                "prompt": build_prompt(cfg, date, chunk_messages, message_start),
            }
        )

    prompt_data = [
        {
            "chunk": item["chunk"],
            "message_start": item["message_start"],
            "message_end": item["message_end"],
            "prompt": item["prompt"],
        }
        for item in chunk_specs
    ]
    prompt: Any = prompt_data[0]["prompt"] if len(prompt_data) == 1 else prompt_data

    if dry_run:
        raw_events = []
        raw_model = None
    else:
        AttractionEvent, _DayEvents = build_models()
        client, needs_model = make_instructor_client(cfg)
        response_model = list[AttractionEvent]
        raw_events = []
        raw_model = []
        for item in chunk_specs:
            print(
                f"[{date}] Instructor + Ollama model: {cfg['model']} "
                f"(chunk {item['chunk']}/{len(chunk_specs)}, "
                f"messages {item['message_start']}-{item['message_end']})"
            )
            result = call_instructor(cfg, client, needs_model, response_model, str(item["prompt"]), max_retries)
            chunk_events = fix_nested_mojibake(extract_events_from_model_result(result))
            raw_events.extend(chunk_events)
            raw_model.append(
                {
                    "chunk": item["chunk"],
                    "message_start": item["message_start"],
                    "message_end": item["message_end"],
                    "events_count": len(chunk_events),
                    "events": chunk_events,
                }
            )
    
    events = enrich_stop_events_from_messages(
        supplement_simple_start_events(normalize_events(raw_events, messages, cfg), messages, cfg),
        messages,
        cfg,
    )
    rows, quarantine = build_rows_from_events(date, events, cfg)

    output = {
        "date": date,
        "model": cfg["model"],
        "messages_count": len(messages),
        "messages": messages,
        "prompt": prompt,
        "raw_model": raw_model,
        "raw_events_count": len(raw_events),
        "raw_events": raw_events,
        "events": events,
        "rows_count": len(rows),
        "rows": rows,
        "quarantine_count": len(quarantine),
        "quarantine": quarantine,
    }
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as fh:
        json.dump(output, fh, ensure_ascii=False, indent=2)
        
    print(f"[{date}] Saved new report: {output_path}")
    print(f"[{date}] events: {len(events)}, rows: {len(rows)}")
    return output, (rows, quarantine)


def main() -> int:
    args = parse_args()
    config_path = Path(args.config).resolve()
    base_dir = config_path.parent
    cfg = fix_nested_mojibake(load_config(config_path))

    json_path = resolve_path(base_dir, cfg["json_path"])
    report_dir = resolve_path(base_dir, cfg.get("report_dir", "reports"))
    by_date = read_messages(json_path)

    if args.date:
        if args.date not in by_date:
            available = ", ".join(sorted(by_date))
            print(f"\n[!] ОШИБКА: Дата {args.date} не найдена в файле {json_path}.", file=sys.stderr)
            return 1
        target_dates = [args.date]
        default_output = report_dir / f"{args.date}_instructor_events.json"
    else:
        month = str(args.month)
        target_dates = [date for date in sorted(by_date) if date.startswith(month + "-")]
        if not target_dates:
            available = ", ".join(sorted(by_date))
            print(f"\n[!] ОШИБКА: Месяц {month} не найден в файле {json_path}.", file=sys.stderr)
            return 1
        default_output = report_dir / f"{month}_instructor_events.json"

    output_path = Path(args.output) if args.output else default_output
    if not output_path.is_absolute():
        output_path = base_dir / output_path

    if len(target_dates) == 1:
        _output, result_tuple = run_for_date(
            cfg,
            target_dates[0],
            by_date[target_dates[0]],
            output_path,
            args.dry_run,
            args.max_retries,
            force=args.force,
            reuse=args.reuse,
        )
        all_rows, all_quarantine = result_tuple
    else:
        days: dict[str, dict[str, Any]] = {}
        all_rows: list[dict[str, Any]] = []
        all_quarantine: list[dict[str, Any]] = []
        for index, date in enumerate(target_dates, start=1):
            print(f"[{index}/{len(target_dates)}] Processing {date}")
            day_output_path = report_dir / f"{date}_instructor_events.json"
            output, result_tuple = run_for_date(
                cfg, 
                date, 
                by_date[date], 
                day_output_path, 
                args.dry_run, 
                args.max_retries,
                force=args.force,
                reuse=args.reuse
            )
            rows, quarantine = result_tuple
            days[date] = output
            all_rows.extend(rows)
            all_quarantine.extend(quarantine)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding="utf-8") as fh:
            json.dump(
                {
                    "month": args.month,
                    "model": cfg["model"],
                    "dates": target_dates,
                    "days": days,
                    "rows_count": len(all_rows),
                    "rows": all_rows,
                    "quarantine_count": len(all_quarantine),
                    "quarantine": all_quarantine,
                },
                fh,
                ensure_ascii=False,
                indent=2,
            )
        print(f"Saved month JSON: {output_path}")
        print(f"Rows: {len(all_rows)}, Quarantine: {len(all_quarantine)}")

    if not args.no_workbook and not args.dry_run:
        workbook_path = resolve_path(base_dir, cfg["workbook_path"])
        output_workbook_path = resolve_path(base_dir, cfg.get("stops_input_path", cfg["output_workbook_path"]))

        # If we use a dedicated stops_input_path, we treat it as both source and output for appending
        if "stops_input_path" in cfg:
            workbook_path = resolve_path(base_dir, cfg["stops_input_path"])
            output_workbook_path = workbook_path

        write_rows_for_dates(cfg, all_rows, target_dates, workbook_path, output_workbook_path, quarantine=all_quarantine)
        
        print(f"Workbook rows written for selected dates: {len(all_rows)}")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("Stopped by user.", file=sys.stderr)
        raise SystemExit(130)
