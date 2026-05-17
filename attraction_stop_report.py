from __future__ import annotations

import argparse
import copy
import datetime as dt
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_CONFIG_PATH = "config.json"
SYSTEM_PROMPT = (
    "Ты извлекаешь факты из сообщений. Не используй режим размышлений, "
    "не показывай ход рассуждений, возвращай только запрошенный результат."
)


def load_config(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        cfg = json.load(fh)

    cfg.setdefault("ollama_url", "http://localhost:11434/api/chat")
    cfg.setdefault("model", "qwen3-vl:4b")
    cfg.setdefault("think", False)
    cfg.setdefault("timeout_seconds", 600)
    cfg.setdefault("options", {"temperature": 0.1})
    cfg.setdefault("json_path", "04.2026.json")
    cfg.setdefault("workbook_path", "Ведомость_аттракционов_2026_шаблон.xlsm")
    cfg.setdefault("output_workbook_path", "Ведомость_аттракционов_2026_заполнено.xlsm")
    cfg.setdefault("input_sheet_name", "Ввод_остановок")
    cfg.setdefault("input_table_name", "tblStops")
    cfg.setdefault("park_open", "10:00")
    cfg.setdefault("park_close", "22:00")
    cfg.setdefault("clear_existing_rows", True)
    cfg.setdefault("report_dir", "reports")
    cfg.setdefault("attractions", [])
    return cfg


def resolve_path(base_dir: Path, value: str | Path) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return base_dir / path


def mojibake_score(text: str) -> int:
    markers = (
        "Р”",
        "Рѕ",
        "Р°",
        "Рµ",
        "Рё",
        "Рљ",
        "Рќ",
        "С‚",
        "СЂ",
        "СЃ",
        "СЊ",
        "С‹",
        "СЋ",
        "вЂ",
        "Ð",
        "Ñ",
    )
    return sum(text.count(marker) for marker in markers)


def fix_mojibake(text: str) -> str:
    if not text:
        return text
    original_score = mojibake_score(text)
    if original_score == 0:
        return text
    try:
        fixed = text.encode("cp1251").decode("utf-8")
    except UnicodeError:
        return text
    return fixed if mojibake_score(fixed) < original_score else text


def read_messages(json_path: Path) -> dict[str, list[dict[str, str]]]:
    with json_path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)

    by_date: dict[str, list[dict[str, str]]] = defaultdict(list)
    for item in data.get("messages", []):
        date = str(item.get("messageDate", "")).strip()
        if not date:
            continue
        text = fix_mojibake(str(item.get("text", "")).strip())
        time = str(item.get("time", "")).strip()
        by_date[date].append({"time": time, "text": text})
    return dict(sorted(by_date.items()))


def format_day_messages(messages: list[dict[str, str]]) -> str:
    lines: list[str] = []
    for msg in messages:
        prefix = msg["time"] if msg["time"] else "--:--"
        lines.append(f"{prefix} {msg['text']}".strip())
    return "\n".join(lines)


def format_numbered_day_messages(messages: list[dict[str, str]]) -> str:
    lines: list[str] = []
    for index, msg in enumerate(messages, start=1):
        prefix = msg["time"] if msg["time"] else "--:--"
        lines.append(f"#{index}. {prefix} {msg['text']}".strip())
    return "\n".join(lines)


def attraction_lines(cfg: dict[str, Any]) -> str:
    lines = []
    for item in cfg.get("attractions", []):
        sheet_name = item.get("sheet_name", "")
        full_name = item.get("full_name", "")
        lines.append(full_name or sheet_name)
    return "\n".join(lines)


def attraction_aliases(cfg: dict[str, Any]) -> list[str]:
    aliases: list[str] = []
    for item in cfg.get("attractions", []):
        names = [
            item.get("sheet_name", ""),
            item.get("full_name", ""),
            *item.get("aliases", []),
        ]
        for name in names:
            text = str(name).strip()
            if text:
                aliases.append(text)
            if "/" in text:
                aliases.extend(part.strip() for part in text.split("/") if part.strip())
    return aliases


def is_global_stop_message(text: str) -> bool:
    lowered = text.lower()
    global_markers = (
        "все аттракционы",
        "всех аттракционах",
        "все уличные аттракционы",
        "аттракционы орион",
        "аттракционы вднх",
        "парк закрыт",
        "парк открыт",
        "продажи закрыты",
        "продажи открыты",
        "продажи прошу не открывать",
    )
    specific_foreign_markers = (
        "учебный полетный центр",
        "автошкола ориона",
        "шахта ориониума",
        "театр сказок",
        "дом страхов",
        "воздушный трамвай",
        "кинопанорама",
    )
    if not any(marker in lowered for marker in global_markers):
        return False
    return not any(marker in lowered for marker in specific_foreign_markers)


def format_relevant_numbered_day_messages(messages: list[dict[str, str]], cfg: dict[str, Any]) -> str:
    aliases = [alias.lower() for alias in attraction_aliases(cfg)]
    lines: list[str] = []
    for index, msg in enumerate(messages, start=1):
        text = msg["text"]
        lowered = text.lower()
        has_target = any(alias and alias in lowered for alias in aliases)
        if not has_target and not is_global_stop_message(text):
            continue
        prefix = msg["time"] if msg["time"] else "--:--"
        lines.append(f"#{index}. {prefix} {text}".strip())
        fragments = target_fragments(text, cfg)
        if fragments:
            for fragment in fragments:
                lines.append(f"   фрагмент: {fragment}")
    return "\n".join(lines)


def stage1_fact_lines(messages: list[dict[str, str]], cfg: dict[str, Any]) -> str:
    lines: list[str] = []
    aliases = [alias.lower() for alias in attraction_aliases(cfg)]
    for index, msg in enumerate(messages, start=1):
        text = msg["text"]
        lowered = text.lower()
        time = msg["time"] if msg["time"] else "--:--"
        for fragment in target_fragments(text, cfg):
            target = fragment.split(" ", 1)[0].strip(' "\'.,')
            for item in cfg.get("attractions", []):
                sheet_name = str(item.get("sheet_name", "")).strip()
                if sheet_name and fragment.lower().startswith(sheet_name.lower()):
                    target = sheet_name
                    break
            lines.append(f"source=#{index}; time={time}; target_hint={target}; fragment={fragment}")
        if is_global_stop_message(text) and not any(alias and alias in lowered for alias in aliases):
            lines.append(f"source=#{index}; time={time}; target_hint=Общее; fragment={text}")
    return "\n".join(lines)


def target_fragments(text: str, cfg: dict[str, Any]) -> list[str]:
    text = re.split(r"\bВ смене\s*:", text, maxsplit=1, flags=re.IGNORECASE)[0]
    # Ищем в тексте не только sheet_name, но и алиасы/полное имя.
    # Важно: возвращаем фрагменты, начинающиеся с канонического sheet_name,
    # чтобы дальше логика target_hint и промпты были стабильны даже при опечатках в чате.
    alias_to_sheet: dict[str, str] = {}
    for item in cfg.get("attractions", []):
        sheet_name = str(item.get("sheet_name", "")).strip()
        if not sheet_name:
            continue
        raw_names = [
            sheet_name,
            str(item.get("full_name", "")).strip(),
            *[str(a).strip() for a in (item.get("aliases", []) or [])],
        ]
        expanded: list[str] = []
        for n in raw_names:
            if not n:
                continue
            expanded.append(n)
            if "/" in n:
                expanded.extend(part.strip() for part in n.split("/") if part.strip())
        for alias in expanded:
            if alias:
                alias_to_sheet[alias] = sheet_name

    positions: list[tuple[int, str, str]] = []
    lowered = text.lower()
    for alias, sheet_name in alias_to_sheet.items():
        start = lowered.find(alias.lower())
        if start >= 0:
            positions.append((start, sheet_name, alias))
    positions.sort()
    if not positions:
        return []

    fragments: list[str] = []
    for idx, (start, sheet_name, alias) in enumerate(positions):
        end = positions[idx + 1][0] if idx + 1 < len(positions) else len(text)
        fragment_body = text[start + len(alias) : end]
        fragment = (sheet_name + fragment_body).strip(" .;")
        fragment = re.sub(r"\s+", " ", fragment)
        if len(fragment) > 180:
            fragment = fragment[:180].rsplit(" ", 1)[0]
        fragments.append(fragment)
    return fragments


def render_template(template: str, values: dict[str, str]) -> str:
    rendered = template
    for key, value in values.items():
        rendered = rendered.replace("{" + key + "}", value)
    return rendered


def strip_thinking(text: str) -> str:
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.IGNORECASE | re.DOTALL)
    return text.strip()


def ask_ollama(cfg: dict[str, Any], prompt: str) -> str:
    payload: dict[str, Any] = {
        "model": cfg["model"],
        "stream": False,
        "messages": [
            {
                "role": "system",
                "content": (
                    "Ты извлекаешь факты из сообщений. Не используй режим размышлений, "
                    "не показывай ход рассуждений, возвращай только запрошенный результат."
                ),
            },
            {"role": "user", "content": prompt},
        ],
        "options": cfg.get("options", {}),
    }
    if "think" in cfg:
        payload["think"] = bool(cfg["think"])

    response = requests.post(
        cfg["ollama_url"],
        json=payload,
        timeout=int(cfg.get("timeout_seconds", 600)),
    )
    response.raise_for_status()
    data = response.json()

    if isinstance(data.get("message"), dict):
        content = data["message"].get("content", "")
    else:
        content = data.get("response", "")
    return strip_thinking(str(content))


def ollama_chat_url(cfg: dict[str, Any]) -> str:
    url = str(cfg.get("ollama_chat_url") or "").strip()
    if url:
        return url
    base = str(cfg.get("ollama_url", "http://localhost:11434/api/chat")).strip()
    if base.endswith("/api/generate"):
        return base[: -len("/api/generate")] + "/api/chat"
    return base if base.endswith("/api/chat") else base.rstrip("/") + "/api/chat"


def ask_ollama_chat_stream(cfg: dict[str, Any], prompt: str, label: str = "Response") -> str:
    payload: dict[str, Any] = {
        "model": cfg["model"],
        "stream": True,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ],
        "options": cfg.get("options", {}),
    }
    if "think" in cfg:
        payload["think"] = bool(cfg["think"])

    url = ollama_chat_url(cfg)
    print(f"Ollama URL: {url}")
    print(f"Model: {cfg['model']}")
    print(f"think: {payload.get('think')}")
    print(f"{label}:")

    chunks: list[str] = []
    thinking_chars = 0
    with requests.post(
        url,
        json=payload,
        stream=True,
        timeout=int(cfg.get("timeout_seconds", 600)),
    ) as response:
        response.raise_for_status()
        for line in response.iter_lines(decode_unicode=True):
            if not line:
                continue
            data = json.loads(line)
            if "error" in data:
                raise RuntimeError(data["error"])
            message = data.get("message") or {}
            thinking_chars += len(str(message.get("thinking") or ""))
            chunk = str(message.get("content") or "")
            if chunk:
                print(chunk, end="", flush=True)
                chunks.append(chunk)
            if data.get("done"):
                break

    print()
    if thinking_chars:
        print(f"Warning: Ollama returned {thinking_chars} thinking characters despite think=false.")
    return strip_thinking("".join(chunks))


def get_park_hours(cfg: dict[str, Any], date_str: str) -> tuple[str, str]:
    try:
        # Expected format YYYY-MM-DD
        date_obj = dt.datetime.strptime(date_str, "%Y-%m-%d")
        if date_obj.weekday() < 4: # Mon-Thu
            return str(cfg.get("park_open_mon_thu", "09:00")), str(cfg.get("park_close_mon_thu", "22:00"))
        else: # Fri-Sun
            return str(cfg.get("park_open_fri_sun", "09:00")), str(cfg.get("park_close_fri_sun", "23:00"))
    except Exception:
        return str(cfg.get("park_open", "09:00")), str(cfg.get("park_close", "22:00"))


def build_prompt(
    cfg: dict[str, Any],
    template_key: str,
    date: str,
    messages_text: str,
    daily_report: str = "",
) -> str:
    park_open, park_close = get_park_hours(cfg, date)
    values = {
        "date": date,
        "messages": messages_text,
        "numbered_messages": messages_text,
        "daily_report": daily_report,
        "stage1_candidates": daily_report,
        "stage2_report": daily_report,
        "attractions": attraction_lines(cfg),
        "park_open": park_open,
        "park_close": park_close,
    }
    return render_template(str(cfg[template_key]), values)


def month_stem(json_path: Path) -> str:
    return json_path.stem.replace(" ", "_")


def run_daily_reports(
    cfg: dict[str, Any],
    by_date: dict[str, list[dict[str, str]]],
    report_dir: Path,
    only_date: str | None = None,
    limit_days: int | None = None,
) -> dict[str, str]:
    daily_reports: dict[str, str] = {}
    dates = [date for date in by_date if only_date is None or date == only_date]
    if limit_days is not None:
        dates = dates[:limit_days]

    for index, date in enumerate(dates, start=1):
        print(f"[1/2] Текстовый отчет: {date} ({index}/{len(dates)})")
        messages_text = format_day_messages(by_date[date])
        prompt = build_prompt(cfg, "daily_report_prompt", date, messages_text)
        daily_reports[date] = ask_ollama(cfg, prompt)

    text_path = report_dir / f"{month_stem(Path(cfg['json_path']))}_daily_text_report.txt"
    with text_path.open("w", encoding="utf-8") as fh:
        for date, report in daily_reports.items():
            fh.write(f"===== {date} =====\n")
            fh.write(report.strip())
            fh.write("\n\n")
    print(f"Текстовый отчет сохранен: {text_path}")
    return daily_reports


def extract_json_object(text: str) -> dict[str, Any]:
    cleaned = strip_thinking(text).strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError(f"Ollama did not return a JSON object: {cleaned[:500]}")
    return json.loads(cleaned[start : end + 1])


def run_structured_pass(
    cfg: dict[str, Any],
    by_date: dict[str, list[dict[str, str]]],
    daily_reports: dict[str, str],
    report_dir: Path,
) -> list[dict[str, Any]]:
    all_rows: list[dict[str, Any]] = []
    raw_by_date: dict[str, str] = {}

    for index, (date, daily_report) in enumerate(daily_reports.items(), start=1):
        print(f"[2/2] JSON для Excel: {date} ({index}/{len(daily_reports)})")
        messages_text = format_day_messages(by_date[date])
        prompt = build_prompt(cfg, "fill_table_prompt", date, messages_text, daily_report)
        raw = ask_ollama(cfg, prompt)
        raw_by_date[date] = raw
        parsed = extract_json_object(raw)

        parsed_date = str(parsed.get("date") or date)
        for row in parsed.get("rows", []):
            if not isinstance(row, dict):
                continue
            row["date"] = parsed_date
            all_rows.append(row)

    rows_path = report_dir / f"{month_stem(Path(cfg['json_path']))}_structured_rows.json"
    raw_path = report_dir / f"{month_stem(Path(cfg['json_path']))}_fill_raw.json"
    with rows_path.open("w", encoding="utf-8") as fh:
        json.dump(all_rows, fh, ensure_ascii=False, indent=2)
    with raw_path.open("w", encoding="utf-8") as fh:
        json.dump(raw_by_date, fh, ensure_ascii=False, indent=2)
    print(f"JSON-строки сохранены: {rows_path}")
    return all_rows


def parse_date(value: Any) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return dt.datetime.strptime(str(value), "%Y-%m-%d").date()


def parse_time(value: Any) -> dt.time | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, dt.time):
        return value.replace(second=0, microsecond=0)
    text = str(value).strip()
    if not text:
        return None
    match = re.search(r"(\d{1,2})[:.](\d{2})", text)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2))
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None
    return dt.time(hour, minute)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def attraction_order(cfg: dict[str, Any]) -> dict[str, int]:
    return {item["sheet_name"]: index for index, item in enumerate(cfg.get("attractions", []))}


def normalize_attraction(value: Any, cfg: dict[str, Any]) -> str:
    text = clean_text(value)
    lowered = text.lower()
    for item in cfg.get("attractions", []):
        names = [
            item.get("sheet_name", ""),
            item.get("full_name", ""),
            *item.get("aliases", []),
        ]
        for name in names:
            name_lower = str(name).lower()
            if name_lower and (lowered == name_lower or name_lower in lowered or lowered in name_lower):
                return item["sheet_name"]
    return text


def should_keep_row(row: dict[str, Any]) -> bool:
    reason = clean_text(row.get("reason"))
    note = clean_text(row.get("note"))
    downtime = clean_text(row.get("downtime"))
    stop = clean_text(row.get("stop_time"))
    start = clean_text(row.get("start_time"))
    joined = " ".join([reason, note, downtime]).lower()
    if "работал без остановок" in joined:
        return False
    return any([reason, note, downtime, stop, start])


def prepare_rows(rows: list[dict[str, Any]], cfg: dict[str, Any]) -> list[dict[str, Any]]:
    order = attraction_order(cfg)
    prepared: list[dict[str, Any]] = []
    for row in rows:
        if not should_keep_row(row):
            continue
        item = dict(row)
        item["date"] = parse_date(item.get("date"))
        item["attraction"] = normalize_attraction(item.get("attraction"), cfg)
        item["stop_time"] = parse_time(item.get("stop_time"))
        item["start_time"] = parse_time(item.get("start_time"))
        item["downtime"] = clean_text(item.get("downtime"))
        item["reason"] = clean_text(item.get("reason"))
        item["note"] = clean_text(item.get("note"))
        prepared.append(item)

    prepared.sort(key=lambda r: (r["date"], order.get(r["attraction"], 999), r["stop_time"] or dt.time(0, 0)))
    return prepared


def copy_row_style(ws: Any, source_row: int, target_row: int, max_col: int = 7) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.font:
            target.font = copy.copy(source.font)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)


def write_workbook(cfg: dict[str, Any], rows: list[dict[str, Any]], workbook_path: Path, output_path: Path) -> None:
    wb = load_workbook(workbook_path, keep_vba=True)
    sheet_name = cfg["input_sheet_name"]
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Лист не найден: {sheet_name}. Доступные листы: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]

    template_row = 2 if ws.max_row >= 2 else 1
    template_styles = []
    for col in range(1, 8):
        cell = ws.cell(template_row, col)
        template_styles.append(
            {
                "style": copy.copy(cell._style),
                "number_format": cell.number_format,
                "alignment": copy.copy(cell.alignment),
                "font": copy.copy(cell.font),
                "fill": copy.copy(cell.fill),
                "border": copy.copy(cell.border),
            }
        )

    if cfg.get("clear_existing_rows", True) and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    prepared = prepare_rows(rows, cfg)
    for offset, row in enumerate(prepared, start=2):
        for col in range(1, 8):
            style = template_styles[col - 1]
            cell = ws.cell(offset, col)
            cell._style = copy.copy(style["style"])
            cell.number_format = style["number_format"]
            cell.alignment = copy.copy(style["alignment"])
            cell.font = copy.copy(style["font"])
            cell.fill = copy.copy(style["fill"])
            cell.border = copy.copy(style["border"])

        ws.cell(offset, 1).value = row["date"]
        ws.cell(offset, 2).value = row["attraction"]
        ws.cell(offset, 3).value = row["stop_time"]
        ws.cell(offset, 4).value = row["start_time"]
        if row["stop_time"] and row["start_time"]:
            ws.cell(offset, 5).value = (
                f'=IF(OR(C{offset}="",D{offset}=""),"",'
                f"IF(D{offset}<C{offset},D{offset}+1-C{offset},D{offset}-C{offset}))"
            )
        elif row["downtime"]:
            ws.cell(offset, 5).value = row["downtime"]
        else:
            ws.cell(offset, 5).value = ""
        ws.cell(offset, 6).value = row["reason"]
        ws.cell(offset, 7).value = row["note"]

    last_row = max(2, len(prepared) + 1)
    table_name = cfg.get("input_table_name")
    if table_name and table_name in ws.tables:
        ws.tables[table_name].ref = f"A1:G{last_row}"

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].bestFit = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Заполненная книга сохранена: {output_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build attraction stop reports from MAX JSON via Ollama and fill the Excel XLSM table."
    )
    parser.add_argument("--config", default=DEFAULT_CONFIG_PATH, help="Path to config.json")
    parser.add_argument("--date", help="Process only one date, YYYY-MM-DD")
    parser.add_argument("--limit-days", type=int, help="Process only the first N dates, useful for tests")
    parser.add_argument("--no-workbook", action="store_true", help="Only create text/JSON reports, do not write XLSM")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    config_path = Path(args.config).resolve()
    base_dir = config_path.parent
    cfg = load_config(config_path)

    json_path = resolve_path(base_dir, cfg["json_path"])
    workbook_path = resolve_path(base_dir, cfg["workbook_path"])
    output_path = resolve_path(base_dir, cfg.get("stops_input_path", cfg["output_workbook_path"]))
    
    # If we use a dedicated stops_input_path, we treat it as both source and output
    if "stops_input_path" in cfg:
        workbook_path = resolve_path(base_dir, cfg["stops_input_path"])
        output_path = workbook_path
    report_dir = resolve_path(base_dir, cfg["report_dir"])
    report_dir.mkdir(parents=True, exist_ok=True)

    cfg["json_path"] = str(json_path)
    by_date = read_messages(json_path)
    if args.date and args.date not in by_date:
        raise KeyError(f"Дата {args.date} не найдена в {json_path}")

    daily_reports = run_daily_reports(cfg, by_date, report_dir, args.date, args.limit_days)
    structured_rows = run_structured_pass(cfg, by_date, daily_reports, report_dir)

    if not args.no_workbook:
        write_workbook(cfg, structured_rows, workbook_path, output_path)

    print(f"Готово. Строк для Excel: {len(prepare_rows(structured_rows, cfg))}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("Остановлено пользователем.", file=sys.stderr)
        raise SystemExit(130)
